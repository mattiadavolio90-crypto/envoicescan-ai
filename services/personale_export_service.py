"""Export Excel mensile turni personale — Fase 5 ristrutturazione Personale.

Riceve dati gia' aggregati (nessun accesso DB qui): il chiamante (router
workspace) e' responsabile di calcolare ore/costi con la stessa logica
condivisa (_ore_turno) usata dalla UI, cosi' i numeri nell'export coincidono
sempre con quelli mostrati in app. Stile ripreso da
services/margine_service.py::export_excel_margini (nessun endpoint proprio,
solo riferimento visivo).
"""
from typing import Optional


def export_excel_personale_mensile(
    turni: list[dict],
    dipendenti: list[dict],
    mese: str,
    nome_ristorante: str,
    ore_standard_per_persona: dict,
    ore_extra_per_persona: dict,
    costo_standard_per_persona: dict,
    costo_extra_per_persona: dict,
    costo_assenze_per_persona: dict,
) -> bytes:
    """Genera il file Excel del mese richiesto.

    Foglio "Turni": righe = dipendenti, colonne = giorni del mese, cella =
    orario+ore per un turno lavorato o sigla per stato-giorno (R/F/M).
    Foglio "Riepilogo": una riga per dipendente con ore/costo std+extra+assenze
    e totale, presi 1:1 dai dizionari aggregati passati dal chiamante.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    weekend_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    tot_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    border_thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    anno, mo = (int(x) for x in mese.split("-"))
    import calendar
    import datetime
    n_giorni = calendar.monthrange(anno, mo)[1]
    giorni_iso = [f"{mese}-{d:02d}" for d in range(1, n_giorni + 1)]
    weekend_flags = [datetime.date(anno, mo, d).weekday() >= 5 for d in range(1, n_giorni + 1)]

    dipendenti_ordinati = sorted(dipendenti, key=lambda d: d["nome"])

    # Pivot dipendente|giorno -> riga turno (una sola per cella, stesso criterio
    # della griglia mensile React: prima occorrenza vince).
    cella: dict[tuple[str, str], dict] = {}
    for t in turni:
        if t.get("mensile"):
            continue
        chiave = (t["dipendente_id"], t["data_turno"])
        if chiave not in cella:
            cella[chiave] = t

    SIGLA = {"riposo": "R", "ferie": "F", "malattia": "M"}

    wb = Workbook()

    # ---- FOGLIO TURNI ----
    ws = wb.active
    ws.title = "Turni"
    tot_cols = 1 + n_giorni
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=tot_cols)
    title_cell = ws.cell(row=1, column=1, value=f"TURNI {mese} — {nome_ristorante}")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    dip_header = ws.cell(row=2, column=1, value="Dipendente")
    dip_header.font = header_font
    dip_header.fill = header_fill
    dip_header.alignment = Alignment(horizontal="center", vertical="center")
    dip_header.border = border_thin
    for i, iso in enumerate(giorni_iso):
        col = 2 + i
        cell = ws.cell(row=2, column=col, value=int(iso.split("-")[2]))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
    ws.row_dimensions[2].height = 20

    for r, dip in enumerate(dipendenti_ordinati):
        row = 3 + r
        nome_cell = ws.cell(row=row, column=1, value=dip["nome"])
        nome_cell.font = Font(bold=True, size=10)
        nome_cell.border = border_thin
        nome_cell.alignment = Alignment(horizontal="left", vertical="center")
        for i, iso in enumerate(giorni_iso):
            col = 2 + i
            t = cella.get((dip["id"], iso))
            valore = ""
            if t:
                tipo = t.get("tipo_giorno", "turno")
                if tipo != "turno":
                    valore = SIGLA.get(tipo, "?")
                else:
                    ora = (t.get("ora_inizio") or "")[:5]
                    ore_tot = _ore_turno_locale(t)
                    valore = f"{ora} ({ore_tot:g}h)" if ora else f"{ore_tot:g}h"
            cell = ws.cell(row=row, column=col, value=valore)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_thin
            cell.font = Font(size=9)
            if weekend_flags[i]:
                cell.fill = weekend_fill

    ws.column_dimensions["A"].width = 22
    for i in range(n_giorni):
        ws.column_dimensions[get_column_letter(2 + i)].width = 10

    # ---- FOGLIO RIEPILOGO ----
    ws2 = wb.create_sheet(title="Riepilogo")
    tot_cols2 = 7
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=tot_cols2)
    title2 = ws2.cell(row=1, column=1, value=f"RIEPILOGO {mese} — {nome_ristorante}")
    title2.font = Font(bold=True, size=14, color="FFFFFF")
    title2.fill = title_fill
    title2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    headers2 = ["Dipendente", "Ore std", "Ore extra", "Costo std (€)", "Costo extra (€)", "Costo assenze (€)", "Totale (€)"]
    for c, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=2, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
    ws2.row_dimensions[2].height = 20

    tot_ore_std = tot_ore_ext = tot_costo_std = tot_costo_ext = tot_costo_ass = 0.0
    for r, dip in enumerate(dipendenti_ordinati):
        row = 3 + r
        nome = dip["nome"]
        ore_std = round(ore_standard_per_persona.get(nome, 0.0), 2)
        ore_ext = round(ore_extra_per_persona.get(nome, 0.0), 2)
        costo_std = round(costo_standard_per_persona.get(nome, 0.0), 2)
        costo_ext = round(costo_extra_per_persona.get(nome, 0.0), 2)
        costo_ass = round(costo_assenze_per_persona.get(nome, 0.0), 2)
        totale = round(costo_std + costo_ext + costo_ass, 2)
        tot_ore_std += ore_std
        tot_ore_ext += ore_ext
        tot_costo_std += costo_std
        tot_costo_ext += costo_ext
        tot_costo_ass += costo_ass

        valori = [nome, ore_std, ore_ext, costo_std, costo_ext, costo_ass, totale]
        for c, v in enumerate(valori, start=1):
            cell = ws2.cell(row=row, column=c, value=v)
            cell.border = border_thin
            if c == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                if c >= 4:
                    cell.number_format = "€ #,##0.00"

    riga_tot = 3 + len(dipendenti_ordinati)
    tot_cell = ws2.cell(row=riga_tot, column=1, value="TOTALE")
    tot_cell.font = Font(bold=True, size=10)
    tot_cell.fill = tot_fill
    tot_cell.border = border_thin
    valori_tot = [
        round(tot_ore_std, 2), round(tot_ore_ext, 2),
        round(tot_costo_std, 2), round(tot_costo_ext, 2), round(tot_costo_ass, 2),
        round(tot_costo_std + tot_costo_ext + tot_costo_ass, 2),
    ]
    for c, v in enumerate(valori_tot, start=2):
        cell = ws2.cell(row=riga_tot, column=c, value=v)
        cell.font = Font(bold=True, size=10)
        cell.fill = tot_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="right")
        if c >= 4:
            cell.number_format = "€ #,##0.00"

    ws2.column_dimensions["A"].width = 22
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        ws2.column_dimensions[col_letter].width = 15

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _ore_turno_locale(t: dict) -> float:
    """Copia minimale di _ore_turno per il solo caso 'turno giornaliero' —
    usata nel foglio Turni (cella) dove serve solo il totale ore del giorno,
    non l'intero split ordinario/extra gia' disponibile nei dict aggregati."""
    def slot(inizio: Optional[str], fine: Optional[str]) -> float:
        if not inizio or not fine:
            return 0.0
        try:
            ih, im = int(inizio[:2]), int(inizio[3:5])
            fh, fm = int(fine[:2]), int(fine[3:5])
        except (ValueError, IndexError):
            return 0.0
        minuti = (fh * 60 + fm) - (ih * 60 + im)
        if minuti < 0:
            minuti += 24 * 60
        return minuti / 60.0

    tot = slot(t.get("ora_inizio"), t.get("ora_fine"))
    if t.get("ora_inizio2") and t.get("ora_fine2"):
        tot += slot(t.get("ora_inizio2"), t.get("ora_fine2"))
    try:
        tot += float(t.get("ore_extra") or 0)
    except (TypeError, ValueError):
        pass
    return round(tot, 2)
