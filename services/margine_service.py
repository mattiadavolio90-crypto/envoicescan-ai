"""
Margine Service - Calcolo MOL (Margine Operativo Lordo)

Gestisce:
- Query costi automatici da fatture (aggregati per mese)
- Caricamento/salvataggio dati margini da Supabase
- Calcoli margini e percentuali
- Export Excel formattato
"""

import pandas as pd
import io
import streamlit as st
from datetime import datetime, timezone
from config.logger_setup import get_logger
from config.constants import CATEGORIE_FOOD, CATEGORIE_SPESE_GENERALI, KPI_SOGLIE
from services import get_supabase_client

logger = get_logger('margine_service')

# Nomi mesi in italiano (abbreviati)
MESI_NOMI = ["Gen", "Feb", "Mar", "Apr", "Mag", "Giu",
             "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]


# ============================================
# COSTI AUTOMATICI DA FATTURE
# ============================================

@st.cache_data(ttl=300, show_spinner="Calcolo costi da fatture...")
def calcola_costi_automatici_per_anno(user_id: str, ristorante_id: str, anno: int) -> tuple:
    """
    Calcola costi F&B e Spese Generali aggregati per mese dalle fatture.
    
    Query diretta a Supabase con colonne snake_case.
    I risultati sono cachati per 5 minuti (TTL 300s).
    
    Args:
        user_id: UUID utente
        ristorante_id: UUID ristorante
        anno: Anno di riferimento (es. 2026)
    
    Returns:
        tuple: (costi_fb_mensili, costi_spese_mensili)
               Entrambi dict {mese_int: somma_float}
    """
    try:
        supabase = get_supabase_client()
        
        # Query con paginazione per dataset grandi
        page_size = 1000
        all_data = []
        offset = 0
        
        while True:
            response = supabase.table('fatture') \
                .select('data_documento, totale_riga, categoria') \
                .eq('user_id', user_id) \
                .eq('ristorante_id', ristorante_id) \
                .is_('deleted_at', 'null') \
                .gte('data_documento', f'{anno}-01-01') \
                .lt('data_documento', f'{anno + 1}-01-01') \
                .neq('categoria', 'Da Classificare') \
                .range(offset, offset + page_size - 1) \
                .execute()
            
            if not response.data:
                break
            
            all_data.extend(response.data)
            
            if len(response.data) < page_size:
                break
            offset += page_size
        
        if not all_data:
            logger.info(f"📊 Nessuna fattura trovata per anno {anno}")
            return {}, {}
        
        df = pd.DataFrame(all_data)
        df['data_documento'] = pd.to_datetime(df['data_documento'], errors='coerce')
        df = df.dropna(subset=['data_documento'])
        df['mese'] = df['data_documento'].dt.month
        
        # Assicurati che totale_riga sia numerico
        df['totale_riga'] = pd.to_numeric(df['totale_riga'], errors='coerce').fillna(0)
        
        # Split F&B vs Spese usando costanti dell'app
        df_fb = df[df['categoria'].isin(CATEGORIE_FOOD)]
        df_spese = df[df['categoria'].isin(CATEGORIE_SPESE_GENERALI)]
        
        # Aggrega per mese
        costi_fb_mensili = df_fb.groupby('mese')['totale_riga'].sum().to_dict()
        costi_spese_mensili = df_spese.groupby('mese')['totale_riga'].sum().to_dict()
        
        logger.info(f"📊 Costi auto {anno}: {len(all_data)} righe fatture, "
                     f"F&B in {len(costi_fb_mensili)} mesi, Spese in {len(costi_spese_mensili)} mesi")
        
        return costi_fb_mensili, costi_spese_mensili
        
    except Exception as e:
        logger.exception(f"❌ Errore calcolo costi automatici anno {anno}: {e}")
        return {}, {}


# ============================================
# COSTI PER CATEGORIA (Analisi Avanzate)
# ============================================

@st.cache_data(ttl=300, show_spinner="Caricamento dati per analisi...")
def carica_costi_per_categoria(user_id: str, ristorante_id: str,
                                date_from: str, date_to: str) -> pd.DataFrame:
    """
    Carica fatture F&B raggruppate per categoria e mese per un periodo.
    
    Args:
        user_id: UUID utente
        ristorante_id: UUID ristorante
        date_from: Data inizio periodo (YYYY-MM-DD)
        date_to: Data fine periodo (YYYY-MM-DD)
    
    Returns:
        DataFrame con colonne: categoria, mese, totale
    """
    try:
        supabase = get_supabase_client()
        
        page_size = 1000
        all_data = []
        offset = 0
        
        while True:
            response = supabase.table('fatture') \
                .select('data_documento, totale_riga, categoria') \
                .eq('user_id', user_id) \
                .eq('ristorante_id', ristorante_id) \
                .is_('deleted_at', 'null') \
                .gte('data_documento', date_from) \
                .lte('data_documento', date_to) \
                .neq('categoria', 'Da Classificare') \
                .range(offset, offset + page_size - 1) \
                .execute()
            
            if not response.data:
                break
            all_data.extend(response.data)
            if len(response.data) < page_size:
                break
            offset += page_size
        
        if not all_data:
            return pd.DataFrame(columns=['categoria', 'mese', 'totale'])
        
        df = pd.DataFrame(all_data)
        df['data_documento'] = pd.to_datetime(df['data_documento'], errors='coerce')
        df = df.dropna(subset=['data_documento'])
        df['mese'] = df['data_documento'].dt.month
        df['totale_riga'] = pd.to_numeric(df['totale_riga'], errors='coerce').fillna(0)

        # Escludi righe con totale non positivo (sconti, rettifiche negative)
        df = df[df['totale_riga'] > 0]

        # Filtra solo F&B
        df = df[df['categoria'].isin(CATEGORIE_FOOD)]
        
        # Aggrega per categoria e mese
        result = df.groupby(['categoria', 'mese'])['totale_riga'].sum().reset_index()
        result.columns = ['categoria', 'mese', 'totale']
        
        logger.info(f"📊 Analisi avanzate {date_from} → {date_to}: "
                     f"{len(result)} righe aggregate, {result['categoria'].nunique()} categorie")
        
        return result
        
    except Exception as e:
        logger.exception(f"❌ Errore caricamento costi per categoria: {e}")
        return pd.DataFrame(columns=['categoria', 'mese', 'totale'])


# ============================================
# CARICAMENTO / SALVATAGGIO MARGINI
# ============================================

def carica_margini_anno(user_id: str, ristorante_id: str, anno: int) -> dict:
    """
    Carica dati margini salvati da Supabase per un anno.
    
    Args:
        user_id: UUID utente
        ristorante_id: UUID ristorante
        anno: Anno di riferimento
    
    Returns:
        dict: {mese_int: row_dict} con i dati salvati
    """
    try:
        supabase = get_supabase_client()
        
        response = supabase.table('margini_mensili') \
            .select('mese, fatturato_iva10, fatturato_iva22, altri_ricavi_noiva, altri_costi_fb, altri_costi_spese, costo_dipendenti') \
            .eq('user_id', user_id) \
            .eq('ristorante_id', ristorante_id) \
            .eq('anno', anno) \
            .execute()
        
        dati = {}
        if response.data:
            for row in response.data:
                dati[row['mese']] = row
            logger.info(f"📊 Caricati {len(dati)} mesi margini per anno {anno}")
        
        return dati
        
    except Exception as e:
        logger.exception(f"❌ Errore caricamento margini anno {anno}: {e}")
        return {}


# Mappa centro → colonna DB. BEVERAGE usa fallback legacy su fatturato_bar
_CENTRO_COL_MAP = {
    "FOOD": "fatturato_food",
    "BEVERAGE": "fatturato_beverage",
    "ALCOLICI": "fatturato_alcolici",
    "DOLCI": "fatturato_dolci",
}

_CENTRO_COL_MAP_LEGACY = {
    "FOOD": "fatturato_food",
    "BEVERAGE": "fatturato_bar",
    "ALCOLICI": "fatturato_alcolici",
    "DOLCI": "fatturato_dolci",
}


def _split_value_for_centro(split_euro: dict, centro: str) -> float:
    """Legge il valore del centro supportando la chiave legacy BAR durante il rename."""
    if centro == "BEVERAGE":
        return float(split_euro.get("BEVERAGE", split_euro.get("BAR", 0.0)) or 0.0)
    return float(split_euro.get(centro, 0.0) or 0.0)


def _is_missing_beverage_column_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return "fatturato_beverage" in text and any(token in text for token in ["column", "schema cache", "does not exist", "pgrst"])


def _select_centri_rows(query_builder, col_map: dict):
    return query_builder.select(','.join(col_map.values())).execute()


def _row_to_centri_dict(row: dict, col_map: dict) -> tuple:
    result = {}
    has_data = False
    for centro, col in col_map.items():
        val = float(row.get(col, 0) or 0)
        result[centro] = val
        if val > 0:
            has_data = True
    return result, has_data


def salva_fatturato_centri(user_id: str, ristorante_id: str, anno: int, mese: int,
                           split_euro: dict) -> bool:
    """
    Salva la suddivisione del fatturato per centro per un mese specifico.
    
    Args:
        user_id: UUID utente
        ristorante_id: UUID ristorante
        anno: Anno
        mese: Mese (1-12)
        split_euro: dict {centro_nome: importo_euro} es. {"FOOD": 30000, "BEVERAGE": 5000, ...}
    
    Returns:
        bool: True se riuscito
    """
    try:
        supabase = get_supabase_client()
        
        record = {
            'user_id': user_id,
            'ristorante_id': ristorante_id,
            'anno': anno,
            'mese': mese,
            'updated_at': datetime.now(timezone.utc).isoformat()
        }
        for centro, col in _CENTRO_COL_MAP.items():
            record[col] = _split_value_for_centro(split_euro, centro)

        try:
            supabase.table('margini_mensili') \
                .upsert(record, on_conflict='ristorante_id,anno,mese') \
                .execute()
        except Exception as e:
            if not _is_missing_beverage_column_error(e):
                raise

            legacy_record = {
                'user_id': user_id,
                'ristorante_id': ristorante_id,
                'anno': anno,
                'mese': mese,
                'updated_at': datetime.now(timezone.utc).isoformat()
            }
            for centro, col in _CENTRO_COL_MAP_LEGACY.items():
                legacy_record[col] = _split_value_for_centro(split_euro, centro)

            supabase.table('margini_mensili') \
                .upsert(legacy_record, on_conflict='ristorante_id,anno,mese') \
                .execute()
        
        logger.info(f"✅ Salvato fatturato centri per {anno}/{mese}")
        return True
        
    except Exception as e:
        logger.exception(f"❌ Errore salvataggio fatturato centri {anno}/{mese}: {e}")
        return False


def carica_fatturato_centri_periodo(user_id: str, ristorante_id: str,
                                     data_inizio, data_fine) -> dict:
    """
    Carica e aggrega la suddivisione fatturato per centro su un range di mesi.
    
    Args:
        user_id: UUID utente
        ristorante_id: UUID ristorante
        data_inizio: date - inizio periodo
        data_fine: date - fine periodo
    
    Returns:
        dict: {centro_nome: totale_euro_nel_periodo} o {} se nessun dato
    """
    try:
        supabase = get_supabase_client()
        
        anno_start = data_inizio.year
        anno_end = data_fine.year
        
        totali = {c: 0.0 for c in _CENTRO_COL_MAP}
        has_data = False
        
        for a in range(anno_start, anno_end + 1):
            m_from = data_inizio.month if a == anno_start else 1
            m_to = data_fine.month if a == anno_end else 12
            
            query = supabase.table('margini_mensili') \
                .eq('user_id', user_id) \
                .eq('ristorante_id', ristorante_id) \
                .eq('anno', a) \
                .gte('mese', m_from) \
                .lte('mese', m_to)

            col_map = _CENTRO_COL_MAP
            try:
                response = _select_centri_rows(query, col_map)
            except Exception as e:
                if not _is_missing_beverage_column_error(e):
                    raise
                col_map = _CENTRO_COL_MAP_LEGACY
                response = _select_centri_rows(query, col_map)
            
            if response.data:
                for row in response.data:
                    for centro, col in col_map.items():
                        val = float(row.get(col, 0) or 0)
                        if val > 0:
                            has_data = True
                        totali[centro] += val
        
        if not has_data:
            return {}
        
        return totali
        
    except Exception as e:
        logger.exception(f"❌ Errore caricamento fatturato centri per periodo: {e}")
        return {}


def carica_fatturato_centri_mese(user_id: str, ristorante_id: str,
                                  anno: int, mese: int) -> dict:
    """
    Carica la suddivisione fatturato per centro per un singolo mese.
    
    Returns:
        dict: {centro_nome: euro} o {} se nessun dato
    """
    try:
        supabase = get_supabase_client()
        
        query = supabase.table('margini_mensili') \
            .eq('user_id', user_id) \
            .eq('ristorante_id', ristorante_id) \
            .eq('anno', anno) \
            .eq('mese', mese)

        col_map = _CENTRO_COL_MAP
        try:
            response = _select_centri_rows(query, col_map)
        except Exception as e:
            if not _is_missing_beverage_column_error(e):
                raise
            col_map = _CENTRO_COL_MAP_LEGACY
            response = _select_centri_rows(query, col_map)
        
        if not response.data:
            return {}
        
        row = response.data[0]
        result, has_data = _row_to_centri_dict(row, col_map)
        
        return result if has_data else {}
        
    except Exception as e:
        logger.exception(f"❌ Errore caricamento fatturato centri mese {anno}/{mese}: {e}")
        return {}


def salva_margini_anno(user_id: str, ristorante_id: str, anno: int,
                       df_input: pd.DataFrame, df_risultati: pd.DataFrame) -> bool:
    """
    Salva/aggiorna i margini di tutti i 12 mesi con upsert atomico.
    
    Salva sempre 12 record (anche mesi a zero) per consistenza.
    
    Args:
        user_id: UUID utente
        ristorante_id: UUID ristorante
        anno: Anno di riferimento
        df_input: DataFrame con input utente (12 righe, index 0-11)
        df_risultati: DataFrame con calcoli (13 righe: 12 mesi + totale)
    
    Returns:
        bool: True se salvataggio riuscito
    """
    try:
        supabase = get_supabase_client()
        records = []

        if len(df_risultati) < 12:
            logger.error(f"dfrisultati ha solo {len(df_risultati)} righe, attese 12")
            return False

        for i in range(12):
            mese_num = i + 1
            
            records.append({
                'user_id': user_id,
                'ristorante_id': ristorante_id,
                'anno': anno,
                'mese': mese_num,
                # Input manuale
                'fatturato_iva10': float(df_input.at[i, 'Fatt_IVA10']),
                'fatturato_iva22': float(df_input.at[i, 'Fatt_IVA22']),
                'altri_ricavi_noiva': float(df_input.at[i, 'Altri_Ricavi_NoIVA']),
                'altri_costi_fb': float(df_input.at[i, 'Altri_FB']),
                'altri_costi_spese': float(df_input.at[i, 'Altri_Spese']),
                'costo_dipendenti': float(df_input.at[i, 'Costo_Dipendenti']),
                # Snapshot costi auto
                'costi_fb_auto': float(df_input.at[i, 'Costi_FB_Auto']),
                'costi_spese_auto': float(df_input.at[i, 'Costi_Spese_Auto']),
                # Snapshot calcoli (i primi 12 righe di df_risultati, NO la riga totale)
                'fatturato_netto': float(df_risultati.at[i, 'Fatt_Netto']),
                'costi_fb_totali': float(df_risultati.at[i, 'Costi_FB']),
                'primo_margine': float(df_risultati.at[i, 'Primo_Margine']),
                'mol': float(df_risultati.at[i, 'MOL']),
                'food_cost_perc': float(df_risultati.at[i, 'FB_Perc']),
                'spese_perc': float(df_risultati.at[i, 'Spese_Perc']),
                'personale_perc': float(df_risultati.at[i, 'Pers_Perc']),
                'mol_perc': float(df_risultati.at[i, 'MOL_Perc']),
                'updated_at': datetime.now(timezone.utc).isoformat()
            })
        
        # Upsert atomico su constraint UNIQUE(ristorante_id, anno, mese)
        supabase.table('margini_mensili') \
            .upsert(records, on_conflict='ristorante_id,anno,mese') \
            .execute()
        
        logger.info(f"✅ Salvati 12 mesi margini per anno {anno}")
        return True
        
    except Exception as e:
        logger.exception(f"❌ Errore salvataggio margini anno {anno}: {e}")
        return False


# ============================================
# CALCOLI MARGINI
# ============================================

def calcola_risultati(df_input: pd.DataFrame) -> pd.DataFrame:
    """
    Calcola margini e percentuali per ogni mese + riga totale anno.
    
    Formula:
        Fatturato Netto = (Fatt_IVA10 / 1.10) + (Fatt_IVA22 / 1.22) + Altri_Ricavi_NoIVA
        Costi F&B Tot = Costi_FB_Auto + Altri_FB
        Primo Margine = Fatt_Netto - Costi_FB_Tot
        MOL = Primo_Margine - Spese_Tot - Costo_Dipendenti
        % = (Costo / Fatt_Netto) * 100  (0 se Fatt_Netto = 0)
    
    Args:
        df_input: DataFrame con 12 righe (una per mese), colonne editabili
    
    Returns:
        DataFrame con 13 righe (12 mesi + TOT ANNO) e colonne risultati
    """
    risultati = []
    
    for i in range(12):
        # Fatturato netto (scorporo IVA vendita + altri ricavi no iva)
        fatt_iva10 = float(df_input.at[i, 'Fatt_IVA10'])
        fatt_iva22 = float(df_input.at[i, 'Fatt_IVA22'])
        altri_ricavi_noiva = float(df_input.at[i, 'Altri_Ricavi_NoIVA'])
        fatt_netto = (fatt_iva10 / 1.10) + (fatt_iva22 / 1.22) + altri_ricavi_noiva
        
        # Costi totali (auto + manuali)
        costi_fb_tot = float(df_input.at[i, 'Costi_FB_Auto']) + float(df_input.at[i, 'Altri_FB'])
        costi_spese_tot = float(df_input.at[i, 'Costi_Spese_Auto']) + float(df_input.at[i, 'Altri_Spese'])
        costi_personale = float(df_input.at[i, 'Costo_Dipendenti'])
        
        # Margini
        primo_margine = fatt_netto - costi_fb_tot
        mol = primo_margine - costi_spese_tot - costi_personale
        
        # Percentuali (solo se c'è fatturato > 0)
        if fatt_netto > 0:
            fb_perc = (costi_fb_tot / fatt_netto) * 100
            spese_perc = (costi_spese_tot / fatt_netto) * 100
            pers_perc = (costi_personale / fatt_netto) * 100
            pm_perc = (primo_margine / fatt_netto) * 100
            mol_perc = (mol / fatt_netto) * 100
        else:
            fb_perc = 0.0
            spese_perc = 0.0
            pers_perc = 0.0
            pm_perc = 0.0
            mol_perc = 0.0
        
        risultati.append({
            'Mese': df_input.at[i, 'Mese'],
            'MeseNum': int(df_input.at[i, 'MeseNum']),
            'Fatt_Netto': round(fatt_netto, 2),
            'Costi_FB': round(costi_fb_tot, 2),
            'FB_Perc': round(fb_perc, 2),
            'Primo_Margine': round(primo_margine, 2),
            'PM_Perc': round(pm_perc, 2),
            'Costi_Spese': round(costi_spese_tot, 2),
            'Spese_Perc': round(spese_perc, 2),
            'Costi_Personale': round(costi_personale, 2),
            'Pers_Perc': round(pers_perc, 2),
            'MOL': round(mol, 2),
            'MOL_Perc': round(mol_perc, 2)
        })
    
    df_risultati = pd.DataFrame(risultati)
    
    # Riga totale anno (% ponderate sui totali, non media dei mesi)
    fatt_netto_tot = df_risultati['Fatt_Netto'].sum()
    costi_fb_tot_anno = df_risultati['Costi_FB'].sum()
    costi_spese_tot_anno = df_risultati['Costi_Spese'].sum()
    costi_pers_tot_anno = df_risultati['Costi_Personale'].sum()
    primo_marg_tot = df_risultati['Primo_Margine'].sum()
    mol_tot = df_risultati['MOL'].sum()
    
    if fatt_netto_tot > 0:
        fb_perc_tot = (costi_fb_tot_anno / fatt_netto_tot) * 100
        spese_perc_tot = (costi_spese_tot_anno / fatt_netto_tot) * 100
        pers_perc_tot = (costi_pers_tot_anno / fatt_netto_tot) * 100
        pm_perc_tot = (primo_marg_tot / fatt_netto_tot) * 100
        mol_perc_tot = (mol_tot / fatt_netto_tot) * 100
    else:
        fb_perc_tot = 0.0
        spese_perc_tot = 0.0
        pers_perc_tot = 0.0
        pm_perc_tot = 0.0
        mol_perc_tot = 0.0
    
    totali = {
        'Mese': 'TOT ANNO',
        'MeseNum': 99,
        'Fatt_Netto': round(fatt_netto_tot, 2),
        'Costi_FB': round(costi_fb_tot_anno, 2),
        'FB_Perc': round(fb_perc_tot, 2),
        'Primo_Margine': round(primo_marg_tot, 2),
        'PM_Perc': round(pm_perc_tot, 2),
        'Costi_Spese': round(costi_spese_tot_anno, 2),
        'Spese_Perc': round(spese_perc_tot, 2),
        'Costi_Personale': round(costi_pers_tot_anno, 2),
        'Pers_Perc': round(pers_perc_tot, 2),
        'MOL': round(mol_tot, 2),
        'MOL_Perc': round(mol_perc_tot, 2)
    }
    
    df_risultati = pd.concat([df_risultati, pd.DataFrame([totali])], ignore_index=True)
    
    return df_risultati


# ============================================
# KPI ANNUALI
# ============================================

def calcola_kpi_anno(df_risultati: pd.DataFrame, mesi_filtro: list = None) -> dict:
    """
    Calcola KPI riepilogo anno dai risultati margini.
    
    Filtra solo mesi con Fatt_Netto > 0 (mesi con attività effettiva).
    Tutte le medie sono ARITMETICHE (somma/numero mesi).
    
    Args:
        df_risultati: DataFrame con 13 righe (12 mesi + TOT ANNO da calcola_risultati)
        mesi_filtro: lista di numeri mese (1-12) da includere. Se None, tutti i 12 mesi.
    
    Returns:
        dict con chiavi: mol_medio, fc_medio, mol_perc_medio, fatt_medio, num_mesi
    """
    kpi_zero = {
        'mol_medio': 0.0,
        'fc_medio': 0.0,
        'mol_perc_medio': 0.0,
        'fatt_medio': 0.0,
        'primo_margine_medio': 0.0,
        'primo_margine_perc_media': 0.0,
        'costi_fb_medi': 0.0,
        'spese_gen_medie': 0.0,
        'spese_gen_perc_media': 0.0,
        'personale_medio': 0.0,
        'personale_perc_media': 0.0,
        'num_mesi': 0
    }
    
    try:
        # Escludi riga TOT ANNO, tieni solo mesi con fatturato > 0
        df_mesi = df_risultati[
            (df_risultati['MeseNum'] != 99) &
            (df_risultati['Fatt_Netto'] > 0)
        ]
        
        # Filtra per periodo selezionato
        if mesi_filtro is not None:
            df_mesi = df_mesi[df_mesi['MeseNum'].isin(mesi_filtro)]
        
        num_mesi = len(df_mesi)
        if num_mesi == 0:
            return kpi_zero
        
        # Medie aritmetiche dei valori mensili
        return {
            'mol_medio': round(df_mesi['MOL'].mean(), 2),
            'fc_medio': round(df_mesi['FB_Perc'].mean(), 2),
            'mol_perc_medio': round(df_mesi['MOL_Perc'].mean(), 2),
            'fatt_medio': round(df_mesi['Fatt_Netto'].mean(), 2),
            'primo_margine_medio': round(df_mesi['Primo_Margine'].mean(), 2),
            'primo_margine_perc_media': round(df_mesi['PM_Perc'].mean(), 2),
            'costi_fb_medi': round(df_mesi['Costi_FB'].mean(), 2),
            'spese_gen_medie': round(df_mesi['Costi_Spese'].mean(), 2),
            'spese_gen_perc_media': round(df_mesi['Spese_Perc'].mean(), 2),
            'personale_medio': round(df_mesi['Costi_Personale'].mean(), 2),
            'personale_perc_media': round(df_mesi['Pers_Perc'].mean(), 2),
            'num_mesi': num_mesi
        }
    except Exception as e:
        logger.error(f"Errore calcolo KPI anno: {e}")
        return kpi_zero


# ============================================
# COMMENTI AUTOMATICI KPI
# ============================================

def _valuta_soglia(valore: float, chiave_soglia: str, crescente: bool = True) -> tuple:
    """
    Valuta un valore rispetto alle soglie definite in KPI_SOGLIE.
    
    Args:
        valore: valore percentuale da valutare
        chiave_soglia: chiave in KPI_SOGLIE
        crescente: True se soglie crescenti (food_cost, spese), False per margini
    
    Returns:
        tuple (emoji, commento)
    """
    soglie = KPI_SOGLIE.get(chiave_soglia, [])
    if not soglie:
        return ('ℹ️', 'Nessuna soglia configurata')
    
    if crescente:
        # Per costi: più basso = meglio (food cost, spese generali)
        for soglia_max, emoji, commento in soglie:
            if valore <= soglia_max:
                return (emoji, commento)
    else:
        # Per margini: più alto = meglio (primo margine, MOL)
        for soglia_min, emoji, commento in soglie:
            if valore <= soglia_min:
                return (emoji, commento)
    
    # Fallback: ultima soglia
    return (soglie[-1][1], soglie[-1][2])


def genera_commenti_kpi(kpi: dict, df_risultati, mesi_filtro: list = None) -> list:
    """
    Genera commenti testuali per ogni KPI basati su soglie del settore ristorazione.
    
    Il fatturato viene valutato confrontando ogni mese del periodo con la media
    di periodo: se la variabilità (coefficiente di variazione) è alta, si segnala.
    
    Args:
        kpi: dict da calcola_kpi_anno()
        df_risultati: DataFrame completo risultati (per analisi fatturato per mese)
        mesi_filtro: lista mesi del periodo selezionato
    
    Returns:
        list di dict con chiavi: kpi_nome, emoji, commento, colore
    """
    commenti = []
    
    if kpi.get('num_mesi', 0) == 0:
        return commenti
    
    # Mappa colori per emoji
    colori = {'🟢': '#16a34a', '🟡': '#ca8a04', '🟠': '#ea580c', '🔴': '#dc2626', 'ℹ️': '#2563eb'}
    
    # 1. Food Cost %
    fc = kpi.get('fc_medio', 0.0)
    emoji, testo = _valuta_soglia(fc, 'food_cost', crescente=True)
    commenti.append({'kpi_nome': 'Food Cost', 'percentuale': f'{fc:.1f}%', 'commento': testo, 'emoji': emoji, 'colore': colori.get(emoji, '#6b7280')})
    
    # 2. 1° Margine %
    pm = kpi.get('primo_margine_perc_media', 0.0)
    emoji, testo = _valuta_soglia(pm, 'primo_margine', crescente=False)
    commenti.append({'kpi_nome': '1° Margine', 'percentuale': f'{pm:.1f}%', 'commento': testo, 'emoji': emoji, 'colore': colori.get(emoji, '#6b7280')})
    
    # 3. Spese Generali %
    sg = kpi.get('spese_gen_perc_media', 0.0)
    emoji, testo = _valuta_soglia(sg, 'spese_generali', crescente=True)
    commenti.append({'kpi_nome': 'Spese Generali', 'percentuale': f'{sg:.1f}%', 'commento': testo, 'emoji': emoji, 'colore': colori.get(emoji, '#6b7280')})
    
    # 4. MOL %
    mol = kpi.get('mol_perc_medio', 0.0)
    emoji, testo = _valuta_soglia(mol, 'mol', crescente=False)
    commenti.append({'kpi_nome': 'MOL', 'percentuale': f'{mol:.1f}%', 'commento': testo, 'emoji': emoji, 'colore': colori.get(emoji, '#6b7280')})

    # 5. Personale %
    pers = kpi.get('personale_perc_media', 0.0)
    emoji, testo = _valuta_soglia(pers, 'personale', crescente=True)
    commenti.append({'kpi_nome': 'Costo del lavoro', 'percentuale': f'{pers:.1f}%', 'commento': testo, 'emoji': emoji, 'colore': colori.get(emoji, '#6b7280')})
    
    # 6. Fatturato — confronto con media di periodo
    try:
        df_mesi = df_risultati[
            (df_risultati['MeseNum'] != 99) &
            (df_risultati['Fatt_Netto'] > 0)
        ]
        if mesi_filtro is not None:
            df_mesi = df_mesi[df_mesi['MeseNum'].isin(mesi_filtro)]
        
        if len(df_mesi) >= 2:
            media_periodo = df_mesi['Fatt_Netto'].mean()
            std_periodo = df_mesi['Fatt_Netto'].std()
            
            if media_periodo > 0:
                cv = (std_periodo / media_periodo) * 100  # Coefficiente di variazione %
                emoji, testo = _valuta_soglia(cv, 'fatturato_variabilita', crescente=True)
                
                # Aggiungere dettaglio sui mesi sopra/sotto media
                mesi_sotto = df_mesi[df_mesi['Fatt_Netto'] < media_periodo * 0.85]
                mesi_sopra = df_mesi[df_mesi['Fatt_Netto'] > media_periodo * 1.15]
                
                dettaglio = testo
                if len(mesi_sotto) > 0:
                    nomi_sotto = ', '.join(df_mesi.loc[mesi_sotto.index, 'Mese'].tolist())
                    dettaglio += f' · Sotto media (-15%): {nomi_sotto}'
                if len(mesi_sopra) > 0:
                    nomi_sopra = ', '.join(df_mesi.loc[mesi_sopra.index, 'Mese'].tolist())
                    dettaglio += f' · Sopra media (+15%): {nomi_sopra}'
                
                commenti.append({'kpi_nome': 'Fatturato', 'percentuale': f'CV {cv:.0f}%', 'commento': dettaglio, 'emoji': emoji, 'colore': colori.get(emoji, '#6b7280')})
    except Exception as e:
        logger.warning(f"Errore analisi fatturato: {e}")
    
    return commenti


# ============================================
# TABELLA TRASPOSTA PER DATA_EDITOR
# ============================================

def build_transposed_df(df_input: pd.DataFrame) -> pd.DataFrame:
    """
    Build transposed display DataFrame with € and % columns for each month.
    
    Structure: Voce | Gen € | Gen % | Feb € | Feb % | ... | Dic € | Dic %
    Total: 1 + 12*2 = 25 columns
    
    Percentages show incidence on Fatturato Netto for that month.
    Calculated rows are prefixed with '=' in the Voce column.
    DB-auto rows have '(Fatture)' suffix.
    
    Args:
        df_input: DataFrame with 12 rows (months) and input columns
    
    Returns:
        DataFrame with 12 rows (voci) and 25 columns (Voce + 12 months * 2)
    """
    # Compute derived values and percentages for each month
    calc = []
    for i in range(12):
        fatt_iva10 = float(df_input.at[i, 'Fatt_IVA10'])
        fatt_iva22 = float(df_input.at[i, 'Fatt_IVA22'])
        altri_ricavi_noiva = float(df_input.at[i, 'Altri_Ricavi_NoIVA'])
        fatt_netto = round((fatt_iva10 / 1.10) + (fatt_iva22 / 1.22) + altri_ricavi_noiva, 2)
        
        costi_fb_auto = float(df_input.at[i, 'Costi_FB_Auto'])
        altri_fb = float(df_input.at[i, 'Altri_FB'])
        costi_fb_tot = round(costi_fb_auto + altri_fb, 2)
        
        primo_margine = round(fatt_netto - costi_fb_tot, 2)
        
        costi_spese_auto = float(df_input.at[i, 'Costi_Spese_Auto'])
        altre_spese = float(df_input.at[i, 'Altri_Spese'])
        spese_tot = round(costi_spese_auto + altre_spese, 2)
        
        costo_personale = float(df_input.at[i, 'Costo_Dipendenti'])
        mol = round(primo_margine - spese_tot - costo_personale, 2)
        
        # Percentages (only if fatt_netto > 0)
        if fatt_netto > 0:
            fb_auto_perc = round((costi_fb_auto / fatt_netto) * 100, 1)
            altri_fb_perc = round((altri_fb / fatt_netto) * 100, 1)
            fb_tot_perc = round((costi_fb_tot / fatt_netto) * 100, 1)
            pm_perc = round((primo_margine / fatt_netto) * 100, 1)
            spese_auto_perc = round((costi_spese_auto / fatt_netto) * 100, 1)
            altre_spese_perc = round((altre_spese / fatt_netto) * 100, 1)
            pers_perc = round((costo_personale / fatt_netto) * 100, 1)
            mol_perc = round((mol / fatt_netto) * 100, 1)
        else:
            fb_auto_perc = 0.0
            altri_fb_perc = 0.0
            fb_tot_perc = 0.0
            pm_perc = 0.0
            spese_auto_perc = 0.0
            altre_spese_perc = 0.0
            pers_perc = 0.0
            mol_perc = 0.0
        
        calc.append({
            'fatt_netto': fatt_netto,
            'costi_fb_tot': costi_fb_tot,
            'fb_tot_perc': fb_tot_perc,
            'primo_margine': primo_margine,
            'pm_perc': pm_perc,
            'spese_tot': spese_tot,
            'mol': mol,
            'fb_auto_perc': fb_auto_perc,
            'altri_fb_perc': altri_fb_perc,
            'spese_auto_perc': spese_auto_perc,
            'altre_spese_perc': altre_spese_perc,
            'pers_perc': pers_perc,
            'mol_perc': mol_perc,
        })
    
    # Define rows: (label, val_getter, perc_getter (or None))
    voci_def = [
        ('Fatt. IVA 10%',            lambda i: float(df_input.at[i, 'Fatt_IVA10']),           None),
        ('Fatt. IVA 22%',            lambda i: float(df_input.at[i, 'Fatt_IVA22']),           None),
        ('Altri ricavi (no iva)',    lambda i: float(df_input.at[i, 'Altri_Ricavi_NoIVA']),   None),
        ('= Fatturato Netto',        lambda i: calc[i]['fatt_netto'],                          None),
        ('Costi F&B (da Fatture)',   lambda i: float(df_input.at[i, 'Costi_FB_Auto']),        lambda i: calc[i]['fb_auto_perc']),
        ('Altri Costi F&B',          lambda i: float(df_input.at[i, 'Altri_FB']),             lambda i: calc[i]['altri_fb_perc']),
        ('= Costi F&B Totali',       lambda i: calc[i]['costi_fb_tot'],                       lambda i: calc[i]['fb_tot_perc']),
        ('= 1° Margine',             lambda i: calc[i]['primo_margine'],                      lambda i: calc[i]['pm_perc']),
        ('Spese Gen. (da Fatture)',  lambda i: float(df_input.at[i, 'Costi_Spese_Auto']),     lambda i: calc[i]['spese_auto_perc']),
        ('Altre Spese Generali',     lambda i: float(df_input.at[i, 'Altri_Spese']),          lambda i: calc[i]['altre_spese_perc']),
        ('Costo personale Lordo',    lambda i: float(df_input.at[i, 'Costo_Dipendenti']),     lambda i: calc[i]['pers_perc']),
        ('= 2° Margine (MOL)',       lambda i: calc[i]['mol'],                                lambda i: calc[i]['mol_perc']),
    ]
    
    rows = []
    for voce_label, val_getter, perc_getter in voci_def:
        row = {'Voce': voce_label}
        for mese_idx, mese in enumerate(MESI_NOMI):
            row[f'{mese} €'] = val_getter(mese_idx)
            if perc_getter:
                p = perc_getter(mese_idx)
                row[f'{mese} %'] = round(p, 1)
            else:
                row[f'{mese} %'] = None
        rows.append(row)
    
    return pd.DataFrame(rows)


# ============================================
# EXPORT EXCEL
# ============================================

def export_excel_margini(df_risultati: pd.DataFrame, anno: int, nome_ristorante: str, kpi_data: dict = None) -> bytes:
    """
    Genera file Excel formattato con report margini annuale - struttura trasposta.
    
    Layout: righe = voci finanziarie, colonne = mesi con dati + TOT ANNO.
    Ogni mese ha 2 sotto-colonne: € e %.
    
    Formattazione:
    - Titolo: sfondo #1F4E78, testo bianco 14pt bold
    - Header mesi: sfondo #4472C4, testo bianco bold
    - Colonne €: sfondo bianco
    - Colonne %: sfondo grigio chiaro (#F5F5F5)
    - Colonne TOT ANNO: sfondo grigio scuro (#E0E0E0), bold
    - Riga MOL: bold con bordo superiore blu
    
    Args:
        df_risultati: DataFrame risultati (13 righe da calcola_risultati)
        anno: Anno di riferimento
        nome_ristorante: Nome ristorante per titolo
        kpi_data: Dizionario con KPI del periodo selezionato (opzionale)
    
    Returns:
        bytes: Contenuto file Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Margini {anno}"
    
    # ---- STILI ----
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_dark_fill = PatternFill(start_color="2C5AA0", end_color="2C5AA0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    perc_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    tot_val_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    tot_perc_fill = PatternFill(start_color="EBEBEB", end_color="EBEBEB", fill_type="solid")
    voce_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    mol_top_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='medium', color="4472C4"),
        bottom=Side(style='thin')
    )
    
    # ---- DETERMINA MESI DA MOSTRARE ----
    # Mostra SEMPRE tutti i 12 mesi (come nella tabella HTML)
    MESI_ORDINE = ['Gen', 'Feb', 'Mar', 'Apr', 'Mag', 'Giu',
                   'Lug', 'Ago', 'Set', 'Ott', 'Nov', 'Dic']
    mesi_da_mostrare = MESI_ORDINE
    num_mesi = 12
    
    # Totale colonne: 1 (Voce) + num_mesi*2 (€ e %) + 2 (TOT ANNO € e %)
    tot_cols = 1 + num_mesi * 2 + 2
    
    # ---- RIGA 1: TITOLO ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=tot_cols)
    title_cell = ws.cell(row=1, column=1, value=f"REPORT MARGINI {anno} - {nome_ristorante}")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # ---- RIGA 2: HEADER MESI (merged su 2 colonne) ----
    # Colonna A: "Voce" (merged righe 2-3)
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    voce_header = ws.cell(row=2, column=1, value="Voce")
    voce_header.font = header_font
    voce_header.fill = header_fill
    voce_header.alignment = Alignment(horizontal='center', vertical='center')
    voce_header.border = border_thin
    ws.cell(row=3, column=1).border = border_thin
    
    col_offset = 2  # Parti da colonna B
    for mese in mesi_da_mostrare:
        # Merge 2 colonne per ogni mese
        ws.merge_cells(start_row=2, start_column=col_offset, end_row=2, end_column=col_offset + 1)
        mese_cell = ws.cell(row=2, column=col_offset, value=mese)
        mese_cell.font = header_font
        mese_cell.fill = header_fill
        mese_cell.alignment = Alignment(horizontal='center', vertical='center')
        mese_cell.border = border_thin
        ws.cell(row=2, column=col_offset + 1).border = border_thin
        col_offset += 2
    
    # TOT ANNO (merge 2 colonne)
    ws.merge_cells(start_row=2, start_column=col_offset, end_row=2, end_column=col_offset + 1)
    tot_cell = ws.cell(row=2, column=col_offset, value="TOT ANNO")
    tot_cell.font = header_font
    tot_cell.fill = header_dark_fill
    tot_cell.alignment = Alignment(horizontal='center', vertical='center')
    tot_cell.border = border_thin
    ws.cell(row=2, column=col_offset + 1).border = border_thin
    ws.row_dimensions[2].height = 25
    
    # ---- RIGA 3: SUB-HEADER € / % ----
    sub_font = Font(bold=True, color="FFFFFF", size=9)
    sub_fill = PatternFill(start_color="5B8BD4", end_color="5B8BD4", fill_type="solid")
    sub_fill_dark = PatternFill(start_color="3A5F9E", end_color="3A5F9E", fill_type="solid")
    
    col_offset = 2
    for _ in range(num_mesi):
        for sub_label in ['€', '%']:
            cell = ws.cell(row=3, column=col_offset, value=sub_label)
            cell.font = sub_font
            cell.fill = sub_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border_thin
            col_offset += 1
    # Sub-header TOT ANNO
    for sub_label in ['€', '%']:
        cell = ws.cell(row=3, column=col_offset, value=sub_label)
        cell.font = sub_font
        cell.fill = sub_fill_dark
        cell.alignment = Alignment(horizontal='center')
        cell.border = border_thin
        col_offset += 1
    ws.row_dimensions[3].height = 20
    
    # ---- RIGHE 4-9: DATI (6 voci) ----
    voci = [
        ('Fatturato Netto',    'Fatt_Netto',       None,         False),
        ('Costi F&B',          'Costi_FB',         'FB_Perc',    False),
        ('1° Margine',         'Primo_Margine',    'PM_Perc',    False),
        ('Spese Generali',     'Costi_Spese',      'Spese_Perc', False),
        ('Personale',          'Costi_Personale',  'Pers_Perc',  False),
        ('2° Margine (MOL)',   'MOL',              'MOL_Perc',   True),
    ]
    
    totale_row = df_risultati[df_risultati['MeseNum'] == 99].iloc[0]
    
    for voce_idx, (voce_label, col_val, col_perc, is_mol) in enumerate(voci):
        excel_row = 4 + voce_idx
        use_border = mol_top_border if is_mol else border_thin
        
        # Colonna A: nome voce
        voce_cell = ws.cell(row=excel_row, column=1, value=voce_label)
        voce_cell.font = Font(bold=True, size=10)
        voce_cell.fill = voce_fill
        voce_cell.alignment = Alignment(horizontal='left', vertical='center')
        voce_cell.border = use_border
        
        col_offset = 2
        # Celle per ogni mese
        for mese in mesi_da_mostrare:
            mese_row_match = df_risultati[df_risultati['Mese'] == mese]
            if len(mese_row_match) > 0:
                mese_row = mese_row_match.iloc[0]
                val = round(mese_row[col_val], 2)
                perc_val = round(mese_row[col_perc], 1) if col_perc else None
            else:
                # Mese non presente (non dovrebbe mai accadere)
                val = 0.0
                perc_val = None
            
            # Cella € (valore)
            val_cell = ws.cell(row=excel_row, column=col_offset, value=val)
            val_cell.number_format = '€ #,##0'
            val_cell.alignment = Alignment(horizontal='right')
            val_cell.border = use_border
            if is_mol:
                val_cell.font = Font(bold=True, size=10)
            
            # Cella % (percentuale)
            col_offset += 1
            if col_perc and perc_val is not None:
                perc_cell = ws.cell(row=excel_row, column=col_offset, value=perc_val / 100)
                perc_cell.number_format = '0.0%'
            else:
                perc_cell = ws.cell(row=excel_row, column=col_offset, value='-')
                perc_cell.alignment = Alignment(horizontal='center')
            perc_cell.fill = perc_fill
            perc_cell.border = use_border
            if is_mol:
                perc_cell.font = Font(bold=True, size=10)
            
            col_offset += 1
        
        # Colonne TOT ANNO (€ e %)
        tot_val_cell = ws.cell(row=excel_row, column=col_offset, value=round(totale_row[col_val], 2))
        tot_val_cell.number_format = '€ #,##0'
        tot_val_cell.font = Font(bold=True, size=10)
        tot_val_cell.fill = tot_val_fill
        tot_val_cell.alignment = Alignment(horizontal='right')
        tot_val_cell.border = use_border
        
        col_offset += 1
        if col_perc:
            tot_perc_val = round(totale_row[col_perc], 1)
            tot_perc_cell = ws.cell(row=excel_row, column=col_offset, value=tot_perc_val / 100)
            tot_perc_cell.number_format = '0.0%'
        else:
            tot_perc_cell = ws.cell(row=excel_row, column=col_offset, value='-')
            tot_perc_cell.alignment = Alignment(horizontal='center')
        tot_perc_cell.font = Font(bold=True, size=10)
        tot_perc_cell.fill = tot_perc_fill
        tot_perc_cell.border = use_border
    
    # ---- AUTO-FIT COLONNE ----
    ws.column_dimensions['A'].width = 18  # Voce
    for col_idx in range(2, tot_cols + 1):
        col_letter = get_column_letter(col_idx)
        # Colonne € più larghe, % più strette
        is_euro_col = (col_idx - 2) % 2 == 0  # B, D, F, ...
        ws.column_dimensions[col_letter].width = 14 if is_euro_col else 8
    
    # ---- FOGLIO KPI (SE FORNITO) ----
    if kpi_data:
        ws_kpi = wb.create_sheet(title="KPI Periodo")
        
        # Titolo
        ws_kpi.merge_cells('A1:D1')
        title_kpi = ws_kpi['A1']
        title_kpi.value = f"KPI - {kpi_data['periodo']} ({kpi_data['num_mesi']} mesi)"
        title_kpi.font = Font(bold=True, size=14, color="FFFFFF")
        title_kpi.fill = title_fill
        title_kpi.alignment = Alignment(horizontal='center', vertical='center')
        ws_kpi.row_dimensions[1].height = 30
        
        # Header
        headers = ['KPI', 'Valore €', 'Incidenza %', '']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_kpi.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
        
        # Dati KPI
        kpi_rows = [
            ('Fatturato Totale', kpi_data.get('fatt_totale', 0.0), None),
            ('Fatturato Medio Mensile', kpi_data['fatt_medio'], None),
            ('Food Cost', kpi_data['costi_fb'], kpi_data['fc_perc']),
            ('1° Margine', kpi_data['primo_marg'], kpi_data['primo_marg_perc']),
            ('Spese Generali', kpi_data['spese_gen'], kpi_data['spese_perc']),
            ('Costo del Lavoro', kpi_data.get('personale', 0.0), kpi_data.get('personale_perc')),
            ('2° Margine (MOL)', kpi_data['mol_medio'], kpi_data['mol_perc']),
        ]
        
        row_idx = 3
        for kpi_label, valore, incidenza in kpi_rows:
            # Nome KPI
            cell = ws_kpi.cell(row=row_idx, column=1, value=kpi_label)
            cell.font = Font(bold=True, size=10)
            cell.fill = voce_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border_thin
            
            # Valore €
            cell = ws_kpi.cell(row=row_idx, column=2, value=round(valore, 2))
            cell.number_format = '€ #,##0'
            cell.alignment = Alignment(horizontal='right')
            cell.border = border_thin
            
            # Incidenza %
            if incidenza is not None:
                cell = ws_kpi.cell(row=row_idx, column=3, value=round(incidenza, 1) / 100)
                cell.number_format = '0.0%'
            else:
                cell = ws_kpi.cell(row=row_idx, column=3, value='-')
                cell.alignment = Alignment(horizontal='center')
            cell.border = border_thin
            
            row_idx += 1
        
        # Auto-fit colonne KPI
        ws_kpi.column_dimensions['A'].width = 25
        ws_kpi.column_dimensions['B'].width = 15
        ws_kpi.column_dimensions['C'].width = 12
        ws_kpi.column_dimensions['D'].width = 5
    
    # ---- SALVA IN BUFFER ----
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
