"""
Test per l'inserimento mensile dei turni personale (totali da busta paga).

Copre la logica lato worker introdotta con la modalità Mensile:
- _ore_turno: per le righe mensili ritorna ore_dichiarate, non calcola dagli orari
- ws_personale_list: aggrega il costo delle righe mensili dal lordo reale (non da tariffa)
- guardia di esclusività giornaliero/mensile per dipendente/mese (HTTP 409)
- validazioni del POST mensile
- get_costo_personale_da_turni (margini): le righe mensili usano il lordo reale

Supabase e auth sono mockati, coerentemente con gli altri test del modulo.
"""

import pytest
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import services.fastapi_worker as worker
import services.routers.margini as margini
import services.routers.workspace as workspace


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _query_mock(execute_data=None):
    """Mock chain del client Supabase (select/eq/gte/lte/order/limit/insert/update/execute)."""
    q = MagicMock()
    for m in ("select", "eq", "gte", "lte", "order", "limit", "insert", "update", "delete", "single", "in_"):
        getattr(q, m).return_value = q
    q.execute.return_value = SimpleNamespace(data=execute_data or [])
    return q


def _patch_workspace(table_side_effect):
    """Patcha auth + supabase + ristorante sul router workspace.

    table_side_effect: callable(table_name) -> query mock, così endpoint che
    interrogano più tabelle/condizioni possono restituire dati diversi."""
    client = MagicMock()
    client.table.side_effect = table_side_effect
    return patch.multiple(
        workspace,
        _resolve_user_from_token=MagicMock(return_value={"id": "user-1"}),
        _get_supabase_client=MagicMock(return_value=client),
        _get_ristorante_id_for_user=MagicMock(return_value="rist-1"),
    ), client


# ---------------------------------------------------------------------------
# _ore_turno — righe mensili vs giornaliere
# ---------------------------------------------------------------------------

class TestOreTurno:

    def test_riga_mensile_usa_ore_dichiarate(self):
        t = {"mensile": True, "ore_dichiarate": 168, "ora_inizio": "00:00", "ora_fine": "00:00"}
        assert worker._ore_turno(t) == 168.0

    def test_riga_mensile_ore_dichiarate_none(self):
        t = {"mensile": True, "ore_dichiarate": None}
        assert worker._ore_turno(t) == 0.0

    def test_riga_giornaliera_calcola_da_orari(self):
        t = {"mensile": False, "ora_inizio": "09:00", "ora_fine": "17:00"}
        assert worker._ore_turno(t) == 8.0

    def test_riga_giornaliera_default_senza_flag(self):
        t = {"ora_inizio": "09:00", "ora_fine": "13:30"}
        assert worker._ore_turno(t) == 4.5

    def test_giornaliero_extra_aggiuntive_al_totale(self):
        # Nuova semantica: ore extra in PIU' rispetto all'orario.
        # 9-17 (8h) + 2h extra = 10h totali.
        t = {"mensile": False, "ora_inizio": "09:00", "ora_fine": "17:00", "ore_extra": 2}
        assert worker._ore_turno(t) == 10.0

    def test_giornaliero_extra_con_slot_spezzato(self):
        # 9-13 (4h) + 18-22 (4h) = 8h orari, + 1.5h extra = 9.5h.
        t = {
            "mensile": False, "ora_inizio": "09:00", "ora_fine": "13:00",
            "ora_inizio2": "18:00", "ora_fine2": "22:00", "ore_extra": 1.5,
        }
        assert worker._ore_turno(t) == 9.5


# ---------------------------------------------------------------------------
# ws_personale_list — aggregazione costi righe mensili
# ---------------------------------------------------------------------------

class TestPersonaleListMensile:
    """ws_personale_list aggrega per dipendente_id ma espone le chiavi per nome
    corrente (join con dipendenti) — coerente col contratto storico verso il
    frontend, anche se la chiave interna è cambiata da Fase 0."""

    def test_costo_mensile_da_lordo(self):
        # Una riga mensile: lordo 1850, di cui 120 extra → ordinario 1730, extra 120.
        riga = {
            "id": "m1", "dipendente_id": "dip-mario", "data_turno": "2026-06-01",
            "mensile": True, "ore_dichiarate": 168, "ore_extra": 8,
            "lordo_mensile": 1850.0, "importo_extra": 120.0,
            "costo_orario": None, "costo_orario_extra": None,
        }
        turni_q = _query_mock([riga])
        dipendenti_q = _query_mock([{"id": "dip-mario", "nome": "Mario"}])
        storico_q = _query_mock([{"dipendente_id": "dip-mario", "costo_orario": None, "costo_orario_extra": None, "data_turno": "2026-06-01"}])
        attivi_q = _query_mock([{"id": "dip-mario", "nome": "Mario", "costo_orario_default": None}])
        calls = {"n": 0}
        def side_effect(_name):
            calls["n"] += 1
            return {1: turni_q, 2: dipendenti_q, 3: storico_q}.get(calls["n"], attivi_q)
        ctx, _ = _patch_workspace(side_effect)
        with ctx:
            res = workspace.ws_personale_list(da="2026-06-01", a="2026-06-30", mensile=True, authorization="Bearer x")
        assert res["monte_ore"]["Mario"] == 168.0
        assert res["ore_extra_per_persona"]["Mario"] == 8.0
        assert res["costo_standard_per_persona"]["Mario"] == 1730.0
        assert res["costo_extra_per_persona"]["Mario"] == 120.0
        assert res["costo_totale"] == 1850.0

    def test_costo_mensile_senza_extra(self):
        riga = {
            "id": "m2", "dipendente_id": "dip-anna", "data_turno": "2026-06-01",
            "mensile": True, "ore_dichiarate": 160, "ore_extra": None,
            "lordo_mensile": 1600.0, "importo_extra": None,
            "costo_orario": None, "costo_orario_extra": None,
        }
        turni_q = _query_mock([riga])
        dipendenti_q = _query_mock([{"id": "dip-anna", "nome": "Anna"}])
        storico_q = _query_mock([{"dipendente_id": "dip-anna", "costo_orario": None, "costo_orario_extra": None, "data_turno": "2026-06-01"}])
        attivi_q = _query_mock([{"id": "dip-anna", "nome": "Anna", "costo_orario_default": None}])
        calls = {"n": 0}
        def side_effect(_name):
            calls["n"] += 1
            return {1: turni_q, 2: dipendenti_q, 3: storico_q}.get(calls["n"], attivi_q)
        ctx, _ = _patch_workspace(side_effect)
        with ctx:
            res = workspace.ws_personale_list(da="2026-06-01", a="2026-06-30", mensile=True, authorization="Bearer x")
        assert res["costo_standard_per_persona"]["Anna"] == 1600.0
        assert res["costo_extra_per_persona"].get("Anna", 0) == 0
        assert res["costo_totale"] == 1600.0


# ---------------------------------------------------------------------------
# Guardia esclusività — POST giornaliero
# ---------------------------------------------------------------------------

class TestEsclusivitaGiornaliero:

    def test_blocca_se_esiste_riga_mensile(self):
        # 1ª query (_dipendente_esiste) trova il dipendente, 2ª
        # (_esiste_riga_mese mensile=True) trova una riga → POST giornaliero respinto.
        dip_q = _query_mock([{"id": "dip-mario"}])
        esiste_q = _query_mock([{"id": "m1"}])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            return dip_q if calls["n"] == 1 else esiste_q
        ctx, _ = _patch_workspace(side_effect)
        body = workspace.NuovoTurnoBody(
            dipendente_id="dip-mario", data_turno="2026-06-10", ora_inizio="09:00", ora_fine="17:00"
        )
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_personale_crea(body=body, authorization="Bearer x")
        assert exc.value.status_code == 409
        assert "mensile" in exc.value.detail.lower()

    def test_ok_se_nessuna_riga_mensile(self):
        dip_q = _query_mock([{"id": "dip-mario"}])
        esiste_q = _query_mock([])           # nessuna riga mensile
        insert_q = _query_mock([{"id": "new", "dipendente_id": "dip-mario"}])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            return {1: dip_q, 2: esiste_q}.get(calls["n"], insert_q)
        ctx, _ = _patch_workspace(side_effect)
        body = workspace.NuovoTurnoBody(
            dipendente_id="dip-mario", data_turno="2026-06-10", ora_inizio="09:00", ora_fine="17:00"
        )
        with ctx:
            res = workspace.ws_personale_crea(body=body, authorization="Bearer x")
        assert res == {"id": "new", "dipendente_id": "dip-mario"}


# ---------------------------------------------------------------------------
# Guardia esclusività + validazioni — POST mensile
# ---------------------------------------------------------------------------

class TestPostMensile:
    """ws_personale_crea_mensile: 1ª query _dipendente_esiste, poi (dopo le
    validazioni numeriche) _esiste_riga_mese mensile=False, poi mensile=True."""

    def _body(self, **kw):
        base = dict(dipendente_id="dip-mario", mese="2026-06", ore_totali=168, lordo=1850.0)
        base.update(kw)
        return workspace.TurnoMensileBody(**base)

    def _dip_ok(self):
        return _query_mock([{"id": "dip-mario"}])

    def test_blocca_se_esistono_turni_giornalieri(self):
        # 1ª query (_dipendente_esiste) trova il dipendente, 2ª (_esiste_riga_mese
        # mensile=False) trova turni → 409.
        dip_q = self._dip_ok()
        giorn_q = _query_mock([{"id": "g1"}])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            return dip_q if calls["n"] == 1 else giorn_q
        ctx, _ = _patch_workspace(side_effect)
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_personale_crea_mensile(body=self._body(), authorization="Bearer x")
        assert exc.value.status_code == 409
        assert "giornalieri" in exc.value.detail.lower()

    def test_blocca_se_esiste_gia_mensile(self):
        # 1ª query (dipendente) ok, 2ª (giornalieri) vuota, 3ª (mensile) trova → 409.
        dip_q = self._dip_ok()
        vuota = _query_mock([])
        mensile_q = _query_mock([{"id": "m1"}])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            return {1: dip_q, 2: vuota}.get(calls["n"], mensile_q)
        ctx, _ = _patch_workspace(side_effect)
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_personale_crea_mensile(body=self._body(), authorization="Bearer x")
        assert exc.value.status_code == 409
        assert "già un inserimento mensile" in exc.value.detail.lower()

    def test_ore_extra_oltre_totali(self):
        dip_q = self._dip_ok()
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            return dip_q if calls["n"] == 1 else _query_mock([])
        ctx, _ = _patch_workspace(side_effect)
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_personale_crea_mensile(body=self._body(ore_extra=200), authorization="Bearer x")
        assert exc.value.status_code == 400

    def test_importo_extra_oltre_lordo(self):
        dip_q = self._dip_ok()
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            return dip_q if calls["n"] == 1 else _query_mock([])
        ctx, _ = _patch_workspace(side_effect)
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_personale_crea_mensile(body=self._body(importo_extra=9999), authorization="Bearer x")
        assert exc.value.status_code == 400

    def test_tutto_zero_respinto(self):
        dip_q = self._dip_ok()
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            return dip_q if calls["n"] == 1 else _query_mock([])
        ctx, _ = _patch_workspace(side_effect)
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_personale_crea_mensile(body=self._body(ore_totali=0, lordo=0), authorization="Bearer x")
        assert exc.value.status_code == 400

    def test_creazione_valida_payload(self):
        # dipendente ok, due _esiste_riga_mese vuote, poi insert.
        dip_q = self._dip_ok()
        vuota1 = _query_mock([])
        vuota2 = _query_mock([])
        insert_q = _query_mock([{"id": "new"}])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            return {1: dip_q, 2: vuota1, 3: vuota2}.get(calls["n"], insert_q)
        ctx, _ = _patch_workspace(side_effect)
        with ctx:
            res = workspace.ws_personale_crea_mensile(
                body=self._body(ore_totali=168, lordo=1850.0, ore_extra=8, importo_extra=120.0, note="  giugno  "),
                authorization="Bearer x",
            )
        assert res == {"id": "new"}
        payload = insert_q.insert.call_args[0][0]
        assert payload["mensile"] is True
        assert payload["data_turno"] == "2026-06-01"       # 1° del mese
        assert payload["ora_inizio"] == "00:00"
        assert payload["ore_dichiarate"] == 168.0
        assert payload["lordo_mensile"] == 1850.0
        assert payload["ore_extra"] == 8.0
        assert payload["importo_extra"] == 120.0
        assert payload["dipendente_id"] == "dip-mario"


# ---------------------------------------------------------------------------
# Margini — costo personale da turni con righe mensili
# ---------------------------------------------------------------------------

class TestMarginiCostoPersonale:

    def _patch_margini(self, turni):
        client = MagicMock()
        client.table.return_value = _query_mock(turni)
        return patch.multiple(
            margini,
            _resolve_user_from_token=MagicMock(return_value={"id": "user-1"}),
            _get_supabase_client=MagicMock(return_value=client),
            _resolve_ristorante_id=MagicMock(return_value="rist-1"),
        )

    def test_riga_mensile_usa_lordo_reale(self):
        turni = [{
            "id": "m1", "nome": "Mario", "data_turno": "2026-06-01",
            "mensile": True, "ore_dichiarate": 168, "ore_extra": 8,
            "lordo_mensile": 1850.0, "importo_extra": 120.0, "costo_orario": None,
        }]
        with self._patch_margini(turni):
            res = margini.get_costo_personale_da_turni(anno=2026, mese=6, authorization="Bearer x")
        assert res["ore_totali"] == 168.0
        assert res["ore_extra"] == 8.0
        assert res["costo_personale_extra"] == 120.0
        assert res["costo_dipendenti"] == 1730.0
        assert res["n_senza_costo"] == 0   # la riga mensile NON conta come senza costo

    def test_riga_giornaliera_senza_extra(self):
        turni = [{
            "id": "g1", "nome": "Anna", "data_turno": "2026-06-10",
            "mensile": False, "ora_inizio": "09:00", "ora_fine": "17:00",
            "ore_extra": 0, "costo_orario": 12.0,
        }]
        with self._patch_margini(turni):
            res = margini.get_costo_personale_da_turni(anno=2026, mese=6, authorization="Bearer x")
        assert res["ore_totali"] == 8.0
        assert res["costo_dipendenti"] == 96.0
        assert res["costo_personale_extra"] == 0.0

    def test_giornaliero_extra_aggiuntive_split_corretto(self):
        # 9-17 (8h ordinarie) + 2h extra = 10h totali. costo std 12, extra 18.
        # ordinario = 8×12 = 96; extra = 2×18 = 36.
        turni = [{
            "id": "g1", "nome": "Anna", "data_turno": "2026-06-10",
            "mensile": False, "ora_inizio": "09:00", "ora_fine": "17:00",
            "ore_extra": 2, "costo_orario": 12.0, "costo_orario_extra": 18.0,
        }]
        with self._patch_margini(turni):
            res = margini.get_costo_personale_da_turni(anno=2026, mese=6, authorization="Bearer x")
        assert res["ore_totali"] == 10.0
        assert res["ore_extra"] == 2.0
        assert res["costo_dipendenti"] == 96.0
        assert res["costo_personale_extra"] == 36.0


# ---------------------------------------------------------------------------
# Fase 1 — rinomina dipendente (PATCH /api/workspace/dipendenti/{id})
# ---------------------------------------------------------------------------

class TestRinominaDipendente:
    """Il rename tocca solo `dipendenti`, mai `turni_personale`, grazie a
    dipendente_id come FK stabile. Copre anche la guardia di unicità tra
    attivi (409) già presente in ws_dipendenti_aggiorna."""

    def _patch_ws(self, omonimo_attivo=None, update_return=None):
        client = MagicMock()
        lookup_result = SimpleNamespace(data=[omonimo_attivo] if omonimo_attivo else [])
        update_result = SimpleNamespace(data=update_return or [])

        def table_side_effect(name):
            if name == "dipendenti":
                q = _query_mock()
                # select/eq/neq/ilike sono tutti chainable sullo stesso mock;
                # solo l'ultima chiamata (ilike per il lookup, update per il salvataggio)
                # importa quale risultato ritorna execute().
                q.select.return_value = q
                q.eq.return_value = q
                q.neq.return_value = q

                def ilike(*a, **kw):
                    q.execute.return_value = lookup_result
                    return q
                q.ilike.side_effect = ilike

                def update(*a, **kw):
                    q.execute.return_value = update_result
                    return q
                q.update.side_effect = update
                return q
            raise AssertionError(f"tabella inattesa: {name}")

        client.table.side_effect = table_side_effect
        return patch.multiple(
            workspace,
            _resolve_user_from_token=MagicMock(return_value={"id": "user-1"}),
            _get_supabase_client=MagicMock(return_value=client),
            _get_ristorante_id_for_user=MagicMock(return_value="rist-1"),
        ), client

    def test_rename_valido_tocca_solo_dipendenti(self):
        aggiornato = {"id": "dip-1", "ristorante_id": "rist-1", "nome": "Marco Rossi", "attivo": True}
        patcher, client = self._patch_ws(omonimo_attivo=None, update_return=[aggiornato])
        with patcher:
            body = workspace.AggiornaDipendenteBody(nome="Marco Rossi")
            res = workspace.ws_dipendenti_aggiorna(dipendente_id="dip-1", body=body, authorization="Bearer x")
        assert res["nome"] == "Marco Rossi"
        tabelle_toccate = {c.args[0] for c in client.table.call_args_list}
        assert tabelle_toccate == {"dipendenti"}

    def test_rename_bloccato_da_omonimo_attivo(self):
        conflitto = {"id": "dip-2", "nome": "Luca Bianchi"}
        patcher, _ = self._patch_ws(omonimo_attivo=conflitto)
        with patcher:
            body = workspace.AggiornaDipendenteBody(nome="Luca Bianchi")
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_dipendenti_aggiorna(dipendente_id="dip-1", body=body, authorization="Bearer x")
        assert exc.value.status_code == 409

    def test_rename_nome_vuoto_respinto(self):
        patcher, _ = self._patch_ws()
        with patcher:
            body = workspace.AggiornaDipendenteBody(nome="   ")
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_dipendenti_aggiorna(dipendente_id="dip-1", body=body, authorization="Bearer x")
        assert exc.value.status_code == 400

    def test_rename_dipendente_inesistente_404(self):
        patcher, _ = self._patch_ws(omonimo_attivo=None, update_return=[])
        with patcher:
            body = workspace.AggiornaDipendenteBody(nome="Nome Nuovo")
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_dipendenti_aggiorna(dipendente_id="dip-inesistente", body=body, authorization="Bearer x")
        assert exc.value.status_code == 404

    def test_solo_costo_orario_non_richiede_guardia_unicita(self):
        aggiornato = {"id": "dip-1", "ristorante_id": "rist-1", "nome": "Marco", "costo_orario_default": 15.0}
        patcher, client = self._patch_ws(update_return=[aggiornato])
        with patcher:
            body = workspace.AggiornaDipendenteBody(costo_orario_default=15.0)
            res = workspace.ws_dipendenti_aggiorna(dipendente_id="dip-1", body=body, authorization="Bearer x")
        assert res["costo_orario_default"] == 15.0


# ---------------------------------------------------------------------------
# Fase 2a — stati giorno espliciti (riposo/ferie/malattia)
# ---------------------------------------------------------------------------

class TestTipoGiorno:
    """tipo_giorno='turno' e' il default che preserva bit-per-bit il
    comportamento pre-esistente (nessuno dei test sopra lo passa mai).
    Le righe di stato azzerano sempre le ore, indipendentemente da eventuali
    orari sporchi lasciati sulla riga."""

    def test_ore_turno_riposo_sempre_zero(self):
        t = {"tipo_giorno": "riposo", "mensile": False, "ora_inizio": "09:00", "ora_fine": "17:00"}
        assert worker._ore_turno(t) == 0.0

    def test_ore_turno_ferie_sempre_zero_anche_con_extra(self):
        t = {"tipo_giorno": "ferie", "mensile": False, "ora_inizio": "09:00", "ora_fine": "17:00", "ore_extra": 2}
        assert worker._ore_turno(t) == 0.0

    def test_ore_turno_malattia_riga_mensile_sempre_zero(self):
        t = {"tipo_giorno": "malattia", "mensile": True, "ore_dichiarate": 168}
        assert worker._ore_turno(t) == 0.0

    def test_ore_turno_default_assente_e_turno_normale(self):
        # Nessuna chiave tipo_giorno in dict: comportamento invariato (Fase 0/1).
        t = {"mensile": False, "ora_inizio": "09:00", "ora_fine": "17:00"}
        assert worker._ore_turno(t) == 8.0


class TestMarginiCostoAssenze:

    def _patch_margini(self, turni):
        client = MagicMock()
        client.table.return_value = _query_mock(turni)
        return patch.multiple(
            margini,
            _resolve_user_from_token=MagicMock(return_value={"id": "user-1"}),
            _get_supabase_client=MagicMock(return_value=client),
            _resolve_ristorante_id=MagicMock(return_value="rist-1"),
        )

    def test_ferie_con_importo_a_carico_isolato_da_costo_dipendenti(self):
        turni = [{
            "id": "f1", "dipendente_id": "dip-anna", "data_turno": "2026-06-10",
            "mensile": False, "tipo_giorno": "ferie", "importo_a_carico": 50.0,
            "ora_inizio": "09:00", "ora_fine": "17:00",
        }]
        with self._patch_margini(turni):
            res = margini.get_costo_personale_da_turni(anno=2026, mese=6, authorization="Bearer x")
        assert res["costo_assenze_a_carico"] == 50.0
        assert res["costo_dipendenti"] == 0.0
        assert res["costo_personale_extra"] == 0.0
        assert res["ore_totali"] == 0.0
        assert res["n_turni"] == 0
        assert res["n_giorni_assenza"] == 1

    def test_riposo_senza_importo_non_incide_su_nulla(self):
        turni = [{
            "id": "r1", "dipendente_id": "dip-anna", "data_turno": "2026-06-11",
            "mensile": False, "tipo_giorno": "riposo", "importo_a_carico": None,
        }]
        with self._patch_margini(turni):
            res = margini.get_costo_personale_da_turni(anno=2026, mese=6, authorization="Bearer x")
        assert res["costo_assenze_a_carico"] == 0.0
        assert res["n_giorni_assenza"] == 1
        assert res["n_senza_costo"] == 0

    def test_turno_lavorato_e_ferie_stesso_mese_si_sommano_correttamente(self):
        turni = [
            {
                "id": "g1", "dipendente_id": "dip-anna", "data_turno": "2026-06-10",
                "mensile": False, "tipo_giorno": "turno",
                "ora_inizio": "09:00", "ora_fine": "17:00", "costo_orario": 12.0, "ore_extra": 0,
            },
            {
                "id": "f1", "dipendente_id": "dip-anna", "data_turno": "2026-06-11",
                "mensile": False, "tipo_giorno": "malattia", "importo_a_carico": 30.0,
            },
        ]
        with self._patch_margini(turni):
            res = margini.get_costo_personale_da_turni(anno=2026, mese=6, authorization="Bearer x")
        assert res["costo_dipendenti"] == 96.0
        assert res["costo_assenze_a_carico"] == 30.0
        assert res["n_turni"] == 1
        assert res["n_giorni_assenza"] == 1
        assert res["ore_totali"] == 8.0


class TestEsclusivitaStatoGiornoMensile:
    """_esiste_turno_lavorato_mese ignora le righe di stato: solo un turno
    tipo_giorno='turno' deve bloccare la coesistenza con una riga mensile."""

    def test_nessun_turno_lavorato_ritorna_false(self):
        q = _query_mock([])  # nessuna riga trovata
        client = MagicMock()
        client.table.return_value = q
        assert workspace._esiste_turno_lavorato_mese(client, "rist-1", "dip-1", "2026-06") is False

    def test_turno_lavorato_esistente_ritorna_true(self):
        q = _query_mock([{"id": "t1"}])
        client = MagicMock()
        client.table.return_value = q
        assert workspace._esiste_turno_lavorato_mese(client, "rist-1", "dip-1", "2026-06") is True


# ---------------------------------------------------------------------------
# Fase 2b — PATCH /stato-giorno (singola riga) e POST /stato-giorno-intervallo
# ---------------------------------------------------------------------------

class TestStatoGiornoPatch:
    """PATCH /api/workspace/personale/{id}/stato-giorno: update diretto sulla
    riga (mensile=False), 404 se non trovata, validazioni su tipo_giorno/importo."""

    def test_cambia_a_riposo_ok(self):
        upd_q = _query_mock([{"id": "t1", "tipo_giorno": "riposo", "importo_a_carico": None}])
        ctx, client = _patch_workspace(lambda _n: upd_q)
        body = workspace.StatoGiornoBody(tipo_giorno="riposo")
        with ctx:
            res = workspace.ws_personale_stato_giorno(turno_id="t1", body=body, authorization="Bearer x")
        assert res["tipo_giorno"] == "riposo"
        upd_q.update.assert_called_once_with({"tipo_giorno": "riposo", "importo_a_carico": None})

    def test_ferie_con_importo_valido(self):
        upd_q = _query_mock([{"id": "t1", "tipo_giorno": "ferie", "importo_a_carico": 50.0}])
        ctx, client = _patch_workspace(lambda _n: upd_q)
        body = workspace.StatoGiornoBody(tipo_giorno="ferie", importo_a_carico=50.0)
        with ctx:
            res = workspace.ws_personale_stato_giorno(turno_id="t1", body=body, authorization="Bearer x")
        assert res["importo_a_carico"] == 50.0

    def test_riga_non_trovata_404(self):
        upd_q = _query_mock([])  # nessuna riga aggiornata (id/ristorante/mensile non combaciano)
        ctx, client = _patch_workspace(lambda _n: upd_q)
        body = workspace.StatoGiornoBody(tipo_giorno="riposo")
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_personale_stato_giorno(turno_id="mancante", body=body, authorization="Bearer x")
        assert exc.value.status_code == 404

    def test_tipo_giorno_non_valido_400(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([]))
        body = SimpleNamespace(tipo_giorno="vacanza", importo_a_carico=None)
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_personale_stato_giorno(turno_id="t1", body=body, authorization="Bearer x")
        assert exc.value.status_code == 400

    def test_importo_su_turno_rifiutato_400(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([]))
        body = workspace.StatoGiornoBody(tipo_giorno="turno", importo_a_carico=10.0)
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_personale_stato_giorno(turno_id="t1", body=body, authorization="Bearer x")
        assert exc.value.status_code == 400

    def test_importo_negativo_rifiutato_400(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([]))
        body = workspace.StatoGiornoBody(tipo_giorno="ferie", importo_a_carico=-5.0)
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_personale_stato_giorno(turno_id="t1", body=body, authorization="Bearer x")
        assert exc.value.status_code == 400


class TestStatoGiornoIntervallo:
    """POST /api/workspace/personale/stato-giorno-intervallo: 1a query
    _dipendente_esiste, 2a select righe esistenti nel range, poi insert/update
    per ogni giorno (mai sovrascrive un turno lavorato esistente)."""

    def _body(self, **kw):
        base = dict(dipendente_id="dip-mario", data_da="2026-06-10", data_a="2026-06-11", tipo_giorno="riposo")
        base.update(kw)
        return workspace.StatoGiornoIntervalloBody(**base)

    def test_dipendente_inesistente_404(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([]))
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_personale_stato_giorno_intervallo(body=self._body(), authorization="Bearer x")
        assert exc.value.status_code == 404

    def test_crea_righe_mancanti_per_ogni_giorno(self):
        dip_q = _query_mock([{"id": "dip-mario"}])
        esistenti_q = _query_mock([])  # nessuna riga nel range
        insert_q = _query_mock([{"id": "new"}])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            if calls["n"] == 1:
                return dip_q
            if calls["n"] == 2:
                return esistenti_q
            return insert_q
        ctx, client = _patch_workspace(side_effect)
        with ctx:
            res = workspace.ws_personale_stato_giorno_intervallo(
                body=self._body(data_da="2026-06-10", data_a="2026-06-11"), authorization="Bearer x"
            )
        assert res["n_creati"] == 2
        assert res["n_aggiornati"] == 0
        assert res["n_saltati_turno_esistente"] == []

    def test_aggiorna_riga_di_stato_esistente(self):
        dip_q = _query_mock([{"id": "dip-mario"}])
        esistenti_q = _query_mock([{"id": "r1", "data_turno": "2026-06-10", "tipo_giorno": "riposo"}])
        update_q = _query_mock([{"id": "r1"}])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            if calls["n"] == 1:
                return dip_q
            if calls["n"] == 2:
                return esistenti_q
            return update_q
        ctx, client = _patch_workspace(side_effect)
        with ctx:
            res = workspace.ws_personale_stato_giorno_intervallo(
                body=self._body(data_da="2026-06-10", data_a="2026-06-10", tipo_giorno="ferie"),
                authorization="Bearer x",
            )
        assert res["n_creati"] == 0
        assert res["n_aggiornati"] == 1
        assert res["n_saltati_turno_esistente"] == []

    def test_salta_giorno_con_turno_lavorato_esistente(self):
        dip_q = _query_mock([{"id": "dip-mario"}])
        esistenti_q = _query_mock([{"id": "t1", "data_turno": "2026-06-10", "tipo_giorno": "turno"}])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            return dip_q if calls["n"] == 1 else esistenti_q
        ctx, client = _patch_workspace(side_effect)
        with ctx:
            res = workspace.ws_personale_stato_giorno_intervallo(
                body=self._body(data_da="2026-06-10", data_a="2026-06-10"), authorization="Bearer x"
            )
        assert res["n_creati"] == 0
        assert res["n_aggiornati"] == 0
        assert res["n_saltati_turno_esistente"] == ["2026-06-10"]

    def test_data_a_precedente_data_da_400(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([{"id": "dip-mario"}]))
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_personale_stato_giorno_intervallo(
                    body=self._body(data_da="2026-06-11", data_a="2026-06-10"), authorization="Bearer x"
                )
        assert exc.value.status_code == 400

    def test_intervallo_troppo_ampio_400(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([{"id": "dip-mario"}]))
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_personale_stato_giorno_intervallo(
                    body=self._body(data_da="2026-01-01", data_a="2028-01-01"), authorization="Bearer x"
                )
        assert exc.value.status_code == 400


class TestRegoleRicorrenti:
    """CRUD /api/workspace/regole-turni: template settimanale per dipendente
    (Fase 3a). 'turno' richiede orari, 'riposo' li rifiuta. Generazione turni
    da regole (POST .../genera) in TestGeneraTurniDaRegole (Fase 3b)."""

    def _body_turno(self, **kw):
        base = dict(
            dipendente_id="dip-mario", giorno_settimana=0, tipo_giorno="turno",
            ora_inizio="09:00", ora_fine="14:00",
        )
        base.update(kw)
        return workspace.NuovaRegolaTurnoBody(**base)

    def _body_riposo(self, **kw):
        base = dict(dipendente_id="dip-mario", giorno_settimana=6, tipo_giorno="riposo")
        base.update(kw)
        return workspace.NuovaRegolaTurnoBody(**base)

    # --- creazione valida --------------------------------------------------

    def test_crea_regola_turno_ok(self):
        dip_q = _query_mock([{"id": "dip-mario"}])
        insert_q = _query_mock([{"id": "r1", "tipo_giorno": "turno", "giorno_settimana": 0}])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            return dip_q if calls["n"] == 1 else insert_q
        ctx, client = _patch_workspace(side_effect)
        with ctx:
            res = workspace.ws_regole_turni_crea(body=self._body_turno(), authorization="Bearer x")
        assert res["tipo_giorno"] == "turno"

    def test_crea_regola_riposo_ok(self):
        dip_q = _query_mock([{"id": "dip-mario"}])
        insert_q = _query_mock([{"id": "r2", "tipo_giorno": "riposo", "giorno_settimana": 6}])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            return dip_q if calls["n"] == 1 else insert_q
        ctx, client = _patch_workspace(side_effect)
        with ctx:
            res = workspace.ws_regole_turni_crea(body=self._body_riposo(), authorization="Bearer x")
        assert res["tipo_giorno"] == "riposo"

    def test_dipendente_inesistente_404(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([]))
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_regole_turni_crea(body=self._body_turno(), authorization="Bearer x")
        assert exc.value.status_code == 404

    # --- validazione orari mancanti/orari su riposo -------------------------

    def test_turno_senza_ora_inizio_400(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([{"id": "dip-mario"}]))
        body = SimpleNamespace(
            dipendente_id="dip-mario", giorno_settimana=0, tipo_giorno="turno",
            ora_inizio=None, ora_fine="14:00", ora_inizio2=None, ora_fine2=None, costo_orario=None,
        )
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_regole_turni_crea(body=body, authorization="Bearer x")
        assert exc.value.status_code == 400

    def test_turno_senza_ora_fine_400(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([{"id": "dip-mario"}]))
        body = SimpleNamespace(
            dipendente_id="dip-mario", giorno_settimana=0, tipo_giorno="turno",
            ora_inizio="09:00", ora_fine=None, ora_inizio2=None, ora_fine2=None, costo_orario=None,
        )
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_regole_turni_crea(body=body, authorization="Bearer x")
        assert exc.value.status_code == 400

    def test_riposo_con_orari_400(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([{"id": "dip-mario"}]))
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_regole_turni_crea(
                    body=self._body_riposo(ora_inizio="09:00"), authorization="Bearer x"
                )
        assert exc.value.status_code == 400

    def test_tipo_giorno_non_valido_400(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([{"id": "dip-mario"}]))
        body = SimpleNamespace(
            dipendente_id="dip-mario", giorno_settimana=0, tipo_giorno="ferie",
            ora_inizio=None, ora_fine=None, ora_inizio2=None, ora_fine2=None, costo_orario=None,
        )
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_regole_turni_crea(body=body, authorization="Bearer x")
        assert exc.value.status_code == 400

    def test_giorno_settimana_fuori_range_400(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([{"id": "dip-mario"}]))
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_regole_turni_crea(
                    body=self._body_turno(giorno_settimana=7), authorization="Bearer x"
                )
        assert exc.value.status_code == 400

    # --- CRUD base -----------------------------------------------------------

    def test_list_regole(self):
        list_q = _query_mock([{"id": "r1", "giorno_settimana": 0}, {"id": "r2", "giorno_settimana": 6}])
        ctx, client = _patch_workspace(lambda _n: list_q)
        with ctx:
            res = workspace.ws_regole_turni_list(dipendente_id=None, attiva=None, authorization="Bearer x")
        assert len(res["regole"]) == 2

    def test_aggiorna_costo_orario(self):
        esistente_q = _query_mock([{
            "id": "r1", "tipo_giorno": "turno", "ora_inizio": "09:00", "ora_fine": "14:00",
            "ora_inizio2": None, "ora_fine2": None,
        }])
        update_q = _query_mock([{"id": "r1", "costo_orario": 12.5}])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            return esistente_q if calls["n"] == 1 else update_q
        ctx, client = _patch_workspace(side_effect)
        body = workspace.AggiornaRegolaTurnoBody(costo_orario=12.5)
        with ctx:
            res = workspace.ws_regole_turni_aggiorna(regola_id="r1", body=body, authorization="Bearer x")
        assert res["costo_orario"] == 12.5

    def test_aggiorna_regola_non_trovata_404(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([]))
        body = workspace.AggiornaRegolaTurnoBody(costo_orario=10.0)
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_regole_turni_aggiorna(regola_id="mancante", body=body, authorization="Bearer x")
        assert exc.value.status_code == 404

    def test_aggiorna_a_riposo_azzera_orari(self):
        esistente_q = _query_mock([{
            "id": "r1", "tipo_giorno": "turno", "ora_inizio": "09:00", "ora_fine": "14:00",
            "ora_inizio2": None, "ora_fine2": None,
        }])
        update_q = _query_mock([{"id": "r1", "tipo_giorno": "riposo", "ora_inizio": None, "ora_fine": None}])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            return esistente_q if calls["n"] == 1 else update_q
        ctx, client = _patch_workspace(side_effect)
        body = workspace.AggiornaRegolaTurnoBody(tipo_giorno="riposo")
        with ctx:
            res = workspace.ws_regole_turni_aggiorna(regola_id="r1", body=body, authorization="Bearer x")
        assert res["tipo_giorno"] == "riposo"
        update_q.update.assert_called_once_with({
            "tipo_giorno": "riposo", "ora_inizio": None, "ora_fine": None,
            "ora_inizio2": None, "ora_fine2": None,
        })

    def test_elimina_regola(self):
        del_q = _query_mock([])
        ctx, client = _patch_workspace(lambda _n: del_q)
        with ctx:
            res = workspace.ws_regole_turni_elimina(regola_id="r1", authorization="Bearer x")
        assert res == {"ok": True}
        del_q.delete.assert_called_once()


# ---------------------------------------------------------------------------
# POST /api/workspace/regole-turni/genera — Fase 3b
# ---------------------------------------------------------------------------

class TestGeneraTurniDaRegole:
    """Genera righe turni_personale dalle regole ricorrenti attive nell'intervallo.
    Le regole 'riposo' non producono righe; niente duplicati su
    (dipendente_id, data_turno) già occupato."""

    def _body(self, **kw):
        base = dict(data_da="2026-08-03", data_a="2026-08-09", dipendente_id=None)  # lun-dom
        base.update(kw)
        return workspace.GeneraTurniDaRegoleBody(**base)

    def test_genera_crea_turni_per_giorno_corrispondente(self):
        # Regola: lunedì (0), dip-mario, 09:00-14:00. Intervallo lun 3/8 - dom 9/8
        # contiene un solo lunedì -> un solo turno creato.
        regole_q = _query_mock([{
            "id": "reg1", "dipendente_id": "dip-mario", "giorno_settimana": 0,
            "tipo_giorno": "turno", "ora_inizio": "09:00", "ora_fine": "14:00",
            "ora_inizio2": None, "ora_fine2": None, "costo_orario": 10.0,
        }])
        esistenti_q = _query_mock([])
        insert_q = _query_mock([])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            if calls["n"] == 1:
                return regole_q
            if calls["n"] == 2:
                return esistenti_q
            return insert_q
        ctx, client = _patch_workspace(side_effect)
        with ctx:
            res = workspace.ws_regole_turni_genera(body=self._body(), authorization="Bearer x")
        assert res == {"ok": True, "n_creati": 1, "n_saltati": 0}
        inserted = insert_q.insert.call_args[0][0]
        assert len(inserted) == 1
        assert inserted[0]["dipendente_id"] == "dip-mario"
        assert inserted[0]["data_turno"] == "2026-08-03"
        assert inserted[0]["ora_inizio"] == "09:00"

    def test_genera_salta_regole_riposo(self):
        # Solo regole tipo_giorno='turno' vengono interrogate (query filtra
        # lato server); qui simuliamo il caso "nessuna regola attiva di tipo turno".
        ctx, client = _patch_workspace(lambda _n: _query_mock([]))
        with ctx:
            res = workspace.ws_regole_turni_genera(body=self._body(), authorization="Bearer x")
        assert res["n_creati"] == 0
        assert "messaggio" in res

    def test_genera_salta_giorno_gia_occupato(self):
        regole_q = _query_mock([{
            "id": "reg1", "dipendente_id": "dip-mario", "giorno_settimana": 0,
            "tipo_giorno": "turno", "ora_inizio": "09:00", "ora_fine": "14:00",
            "ora_inizio2": None, "ora_fine2": None, "costo_orario": None,
        }])
        esistenti_q = _query_mock([{"dipendente_id": "dip-mario", "data_turno": "2026-08-03"}])
        insert_q = _query_mock([])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            if calls["n"] == 1:
                return regole_q
            if calls["n"] == 2:
                return esistenti_q
            return insert_q
        ctx, client = _patch_workspace(side_effect)
        with ctx:
            res = workspace.ws_regole_turni_genera(body=self._body(), authorization="Bearer x")
        assert res == {"ok": True, "n_creati": 0, "n_saltati": 1}
        insert_q.insert.assert_not_called()

    def test_genera_non_salta_stesso_giorno_altro_dipendente(self):
        # Turno esistente per dip-luigi lo stesso giorno non deve bloccare la
        # generazione per dip-mario (chiave composita, non solo data_turno).
        regole_q = _query_mock([{
            "id": "reg1", "dipendente_id": "dip-mario", "giorno_settimana": 0,
            "tipo_giorno": "turno", "ora_inizio": "09:00", "ora_fine": "14:00",
            "ora_inizio2": None, "ora_fine2": None, "costo_orario": None,
        }])
        esistenti_q = _query_mock([{"dipendente_id": "dip-luigi", "data_turno": "2026-08-03"}])
        insert_q = _query_mock([])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            if calls["n"] == 1:
                return regole_q
            if calls["n"] == 2:
                return esistenti_q
            return insert_q
        ctx, client = _patch_workspace(side_effect)
        with ctx:
            res = workspace.ws_regole_turni_genera(body=self._body(), authorization="Bearer x")
        assert res == {"ok": True, "n_creati": 1, "n_saltati": 0}

    def test_dipendente_inesistente_404(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([]))
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_regole_turni_genera(
                    body=self._body(dipendente_id="dip-fantasma"), authorization="Bearer x"
                )
        assert exc.value.status_code == 404

    def test_data_a_precedente_data_da_400(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([]))
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_regole_turni_genera(
                    body=self._body(data_da="2026-08-09", data_a="2026-08-03"), authorization="Bearer x"
                )
        assert exc.value.status_code == 400

    def test_intervallo_troppo_ampio_400(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([]))
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_regole_turni_genera(
                    body=self._body(data_da="2026-01-01", data_a="2026-12-31"), authorization="Bearer x"
                )
        assert exc.value.status_code == 400

    def test_date_non_valide_400(self):
        ctx, client = _patch_workspace(lambda _n: _query_mock([]))
        body = SimpleNamespace(data_da="non-una-data", data_a="2026-08-09", dipendente_id=None)
        with ctx:
            with pytest.raises(worker.HTTPException) as exc:
                workspace.ws_regole_turni_genera(body=body, authorization="Bearer x")
        assert exc.value.status_code == 400


# ---------------------------------------------------------------------------
# POST /api/workspace/personale/copia-settimana — fix chiave composita
# ---------------------------------------------------------------------------

class TestCopiaSettimanaChiaveComposita:
    """Bug preesistente: 'giorni pieni' era chiavizzato solo su data_turno,
    quindi un turno esistente di un dipendente bloccava la copia per TUTTI
    gli altri dipendenti in quel giorno. Fix: chiave (dipendente_id, data_turno)."""

    def _body(self, **kw):
        base = dict(da="2026-08-10", a="2026-08-16")  # lun-dom, settimana successiva
        base.update(kw)
        return workspace.CopiaSettimanaBody(**base)

    def test_non_salta_altro_dipendente_stesso_giorno(self):
        # Sorgente (settimana 3-9/8): dip-mario lunedì 3/8.
        sorgente_q = _query_mock([{
            "dipendente_id": "dip-mario", "data_turno": "2026-08-03",
            "ora_inizio": "09:00", "ora_fine": "14:00",
        }])
        # Destinazione: dip-luigi ha gia' un turno lunedì 10/8 (altro dipendente).
        esistenti_q = _query_mock([{"dipendente_id": "dip-luigi", "data_turno": "2026-08-10"}])
        insert_q = _query_mock([])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            if calls["n"] == 1:
                return sorgente_q
            if calls["n"] == 2:
                return esistenti_q
            return insert_q
        ctx, client = _patch_workspace(side_effect)
        with ctx:
            res = workspace.ws_personale_copia_settimana(body=self._body(), authorization="Bearer x")
        assert res == {"ok": True, "n_copiati": 1, "n_saltati": 0}

    def test_salta_stesso_dipendente_stesso_giorno(self):
        sorgente_q = _query_mock([{
            "dipendente_id": "dip-mario", "data_turno": "2026-08-03",
            "ora_inizio": "09:00", "ora_fine": "14:00",
        }])
        esistenti_q = _query_mock([{"dipendente_id": "dip-mario", "data_turno": "2026-08-10"}])
        insert_q = _query_mock([])
        calls = {"n": 0}
        def side_effect(_n):
            calls["n"] += 1
            if calls["n"] == 1:
                return sorgente_q
            if calls["n"] == 2:
                return esistenti_q
            return insert_q
        ctx, client = _patch_workspace(side_effect)
        with ctx:
            res = workspace.ws_personale_copia_settimana(body=self._body(), authorization="Bearer x")
        assert res == {"ok": True, "n_copiati": 0, "n_saltati": 1}
        insert_q.insert.assert_not_called()


# ---------------------------------------------------------------------------
# Fase 5 — Export Excel mensile
# ---------------------------------------------------------------------------

class TestExportExcelPersonaleMensile:
    """services.personale_export_service.export_excel_personale_mensile:
    verifica che produca un .xlsx valido con i due fogli attesi e che i
    totali del Riepilogo corrispondano ai dati aggregati passati in input."""

    def test_produce_xlsx_valido_con_due_fogli(self):
        from io import BytesIO
        from openpyxl import load_workbook
        from services.personale_export_service import export_excel_personale_mensile

        turni = [
            {"dipendente_id": "d1", "data_turno": "2026-07-01", "ora_inizio": "09:00", "ora_fine": "17:00", "tipo_giorno": "turno"},
            {"dipendente_id": "d1", "data_turno": "2026-07-02", "tipo_giorno": "ferie", "importo_a_carico": 50.0},
            {"dipendente_id": "d2", "data_turno": "2026-07-01", "ora_inizio": "10:00", "ora_fine": "18:00", "ore_extra": 2, "tipo_giorno": "turno"},
        ]
        dipendenti = [{"id": "d1", "nome": "Mario Rossi"}, {"id": "d2", "nome": "Anna Bianchi"}]

        xlsx = export_excel_personale_mensile(
            turni=turni, dipendenti=dipendenti, mese="2026-07", nome_ristorante="Test SRL",
            ore_standard_per_persona={"Mario Rossi": 8.0, "Anna Bianchi": 8.0},
            ore_extra_per_persona={"Anna Bianchi": 2.0},
            costo_standard_per_persona={"Mario Rossi": 80.0, "Anna Bianchi": 80.0},
            costo_extra_per_persona={"Anna Bianchi": 20.0},
            costo_assenze_per_persona={"Mario Rossi": 50.0},
        )
        assert isinstance(xlsx, bytes) and len(xlsx) > 0

        wb = load_workbook(BytesIO(xlsx))
        assert wb.sheetnames == ["Turni", "Riepilogo"]

        ws_turni = wb["Turni"]
        assert ws_turni["A1"].value == "TURNI 2026-07 — Test SRL"
        assert ws_turni["B2"].value == 1  # giorno 1 del mese in colonna B

        ws_riep = wb["Riepilogo"]
        righe = {r[0]: r for r in ws_riep.iter_rows(min_row=3, values_only=True) if r[0]}
        # Mario: 8h std, 80€ std + 50€ assenze = 130€ totale (nessuna extra)
        assert righe["Mario Rossi"][1] == 8.0
        assert righe["Mario Rossi"][3] == 8.0   # ore totali = std + extra
        assert righe["Mario Rossi"][7] == 130.0
        # Anna: 8h std + 2h extra, 80+20 = 100€ totale (nessuna assenza)
        assert righe["Anna Bianchi"][2] == 2.0
        assert righe["Anna Bianchi"][3] == 10.0
        assert righe["Anna Bianchi"][7] == 100.0
        # Riga TOTALE: somma dei due dipendenti
        assert righe["TOTALE"][3] == 18.0
        assert righe["TOTALE"][7] == 230.0

    def test_cella_mostra_orario_inizio_e_fine(self):
        """Il bug segnalato il 4/9/2026: la cella mostrava solo l'ora di
        inizio, quindi dall'export non si poteva sapere quando finiva il
        turno. Deve comparire la fascia intera, spezzato incluso."""
        from io import BytesIO
        from openpyxl import load_workbook
        from services.personale_export_service import export_excel_personale_mensile

        turni = [
            {"dipendente_id": "d1", "data_turno": "2026-07-01", "ora_inizio": "14:00", "ora_fine": "22:00", "tipo_giorno": "turno"},
            {"dipendente_id": "d2", "data_turno": "2026-07-01", "ora_inizio": "08:00", "ora_fine": "14:00",
             "ora_inizio2": "18:00", "ora_fine2": "23:00", "tipo_giorno": "turno"},
        ]
        xlsx = export_excel_personale_mensile(
            turni=turni,
            dipendenti=[{"id": "d1", "nome": "Mario Rossi"}, {"id": "d2", "nome": "Anna Bianchi"}],
            mese="2026-07", nome_ristorante="Test SRL",
            ore_standard_per_persona={}, ore_extra_per_persona={},
            costo_standard_per_persona={}, costo_extra_per_persona={}, costo_assenze_per_persona={},
        )
        ws = load_workbook(BytesIO(xlsx))["Turni"]
        # Anna Bianchi ordina prima di Mario Rossi -> riga 3 / riga 4.
        assert ws.cell(row=4, column=2).value == "14:00-22:00 (8h)"
        assert ws.cell(row=3, column=2).value == "08:00-14:00 / 18:00-23:00 (11h)"

    def test_totale_ore_per_dipendente_e_complessivo_nel_foglio_turni(self):
        """Colonna TOT ORE a fine riga + riga TOTALE ORE in fondo: sommano le
        ore effettivamente in griglia, assenze escluse."""
        from io import BytesIO
        from openpyxl import load_workbook
        from services.personale_export_service import export_excel_personale_mensile

        turni = [
            {"dipendente_id": "d1", "data_turno": "2026-07-01", "ora_inizio": "14:00", "ora_fine": "22:00", "tipo_giorno": "turno"},
            {"dipendente_id": "d1", "data_turno": "2026-07-02", "ora_inizio": "09:00", "ora_fine": "14:00", "tipo_giorno": "turno"},
            {"dipendente_id": "d1", "data_turno": "2026-07-03", "tipo_giorno": "ferie"},
            {"dipendente_id": "d2", "data_turno": "2026-07-01", "ora_inizio": "08:00", "ora_fine": "16:00", "ore_extra": 2, "tipo_giorno": "turno"},
        ]
        xlsx = export_excel_personale_mensile(
            turni=turni,
            dipendenti=[{"id": "d1", "nome": "Mario Rossi"}, {"id": "d2", "nome": "Anna Bianchi"}],
            mese="2026-07", nome_ristorante="Test SRL",
            ore_standard_per_persona={}, ore_extra_per_persona={},
            costo_standard_per_persona={}, costo_extra_per_persona={}, costo_assenze_per_persona={},
        )
        ws = load_workbook(BytesIO(xlsx))["Turni"]
        col_tot = 2 + 31  # luglio: 31 giorni
        assert ws.cell(row=2, column=col_tot).value == "TOT ORE"
        assert ws.cell(row=3, column=1).value == "Anna Bianchi"
        assert ws.cell(row=3, column=col_tot).value == 10.0   # 8h + 2h extra
        assert ws.cell(row=4, column=col_tot).value == 13.0   # 8h + 5h, ferie escluse
        assert ws.cell(row=5, column=1).value == "TOTALE ORE"
        assert ws.cell(row=5, column=col_tot).value == 23.0

    def test_riepilogo_ha_ore_totali_per_dipendente_e_complessive(self):
        """Seconda scheda: colonna Ore totali (std+extra) per riga e nella
        riga TOTALE."""
        from io import BytesIO
        from openpyxl import load_workbook
        from services.personale_export_service import export_excel_personale_mensile

        xlsx = export_excel_personale_mensile(
            turni=[],
            dipendenti=[{"id": "d1", "nome": "Mario Rossi"}, {"id": "d2", "nome": "Anna Bianchi"}],
            mese="2026-07", nome_ristorante="Test SRL",
            ore_standard_per_persona={"Mario Rossi": 160.0, "Anna Bianchi": 120.5},
            ore_extra_per_persona={"Mario Rossi": 12.0},
            costo_standard_per_persona={}, costo_extra_per_persona={}, costo_assenze_per_persona={},
        )
        ws = load_workbook(BytesIO(xlsx))["Riepilogo"]
        intestazioni = [c.value for c in ws[2]]
        assert intestazioni[3] == "Ore totali"
        righe = {r[0]: r for r in ws.iter_rows(min_row=3, values_only=True) if r[0]}
        assert righe["Mario Rossi"][3] == 172.0
        assert righe["Anna Bianchi"][3] == 120.5
        assert righe["TOTALE"][3] == 292.5

    def test_mese_senza_turni_non_solleva_eccezioni(self):
        from services.personale_export_service import export_excel_personale_mensile

        xlsx = export_excel_personale_mensile(
            turni=[], dipendenti=[], mese="2026-07", nome_ristorante="Test SRL",
            ore_standard_per_persona={}, ore_extra_per_persona={},
            costo_standard_per_persona={}, costo_extra_per_persona={}, costo_assenze_per_persona={},
        )
        assert isinstance(xlsx, bytes) and len(xlsx) > 0


class TestEndpointExportMensile:
    """ws_personale_export_mensile: verifica che l'endpoint risponda con un
    Response Excel valido, riusando ws_personale_list per l'aggregazione."""

    def test_ritorna_xlsx_con_content_disposition(self):
        from io import BytesIO
        from openpyxl import load_workbook

        riga = {
            "id": "t1", "dipendente_id": "dip-mario", "data_turno": "2026-07-01",
            "ora_inizio": "09:00", "ora_fine": "17:00", "mensile": False,
            "tipo_giorno": "turno", "costo_orario": 10.0, "costo_orario_extra": None,
            "ore_extra": None,
        }
        turni_q = _query_mock([riga])
        dipendenti_q = _query_mock([{"id": "dip-mario", "nome": "Mario"}])
        storico_q = _query_mock([{"dipendente_id": "dip-mario", "costo_orario": 10.0, "costo_orario_extra": None, "data_turno": "2026-07-01"}])
        attivi_q = _query_mock([{"id": "dip-mario", "nome": "Mario", "costo_orario_default": None}])
        ristorante_q = _query_mock({"nome_ristorante": "Trattoria Test"})
        calls = {"n": 0}
        def side_effect(name):
            calls["n"] += 1
            if name == "ristoranti":
                return ristorante_q
            return {1: turni_q, 2: dipendenti_q, 3: storico_q}.get(calls["n"], attivi_q)
        ctx, _ = _patch_workspace(side_effect)
        with ctx:
            res = workspace.ws_personale_export_mensile(mese="2026-07", authorization="Bearer x")

        assert res.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert res.headers["content-disposition"] == 'attachment; filename="personale_mensile_202607.xlsx"'
        wb = load_workbook(BytesIO(res.body))
        assert wb.sheetnames == ["Turni", "Riepilogo"]
        assert "Trattoria Test" in wb["Turni"]["A1"].value

    def test_mese_non_valido_400(self):
        ctx, _ = _patch_workspace(lambda _n: _query_mock([]))
        with ctx:
            with pytest.raises(Exception) as exc_info:
                workspace.ws_personale_export_mensile(mese="non-valido", authorization="Bearer x")
        assert getattr(exc_info.value, "status_code", None) == 400

    def test_formato_mese_permissivo_respinto_400(self):
        # _mese_bounds da solo accetterebbe '2026-7' (mese senza zero) — la
        # validazione stretta aggiunta per l'export lo respinge comunque,
        # perché il valore finisce nel titolo del foglio e nel filename.
        ctx, _ = _patch_workspace(lambda _n: _query_mock([]))
        with ctx:
            with pytest.raises(Exception) as exc_info:
                workspace.ws_personale_export_mensile(mese="2026-7", authorization="Bearer x")
        assert getattr(exc_info.value, "status_code", None) == 400

    def test_ex_dipendente_incluso_nel_totale(self):
        """Bug trovato in review: ws_personale_list filtra i dipendenti su
        attivo=True, ma turni/dizionari aggregati includono anche chi e' stato
        disattivato durante il mese esportato. Senza il fix quella persona (e
        il suo costo) sparirebbe silenziosamente sia dalla griglia sia dal
        TOTALE del Riepilogo."""
        from io import BytesIO
        from openpyxl import load_workbook

        riga_attivo = {
            "id": "t1", "dipendente_id": "dip-mario", "data_turno": "2026-07-01",
            "ora_inizio": "09:00", "ora_fine": "17:00", "mensile": False,
            "tipo_giorno": "turno", "costo_orario": 10.0, "costo_orario_extra": None,
            "ore_extra": None,
        }
        riga_ex_dipendente = {
            "id": "t2", "dipendente_id": "dip-luigi", "data_turno": "2026-07-05",
            "ora_inizio": "09:00", "ora_fine": "17:00", "mensile": False,
            "tipo_giorno": "turno", "costo_orario": 10.0, "costo_orario_extra": None,
            "ore_extra": None,
        }
        turni_q = _query_mock([riga_attivo, riga_ex_dipendente])
        dipendenti_q = _query_mock([{"id": "dip-mario", "nome": "Mario"}, {"id": "dip-luigi", "nome": "Luigi"}])
        storico_q = _query_mock([
            {"dipendente_id": "dip-mario", "costo_orario": 10.0, "costo_orario_extra": None, "data_turno": "2026-07-01"},
            {"dipendente_id": "dip-luigi", "costo_orario": 10.0, "costo_orario_extra": None, "data_turno": "2026-07-05"},
        ])
        # Solo Mario e' attivo: Luigi e' stato disattivato dopo il turno del 5/7.
        attivi_q = _query_mock([{"id": "dip-mario", "nome": "Mario", "costo_orario_default": None}])
        # Query aggiuntiva del fix: recupero per id degli ex-dipendenti mancanti.
        ex_dipendenti_q = _query_mock([{"id": "dip-luigi", "nome": "Luigi"}])
        ristorante_q = _query_mock({"nome_ristorante": "Trattoria Test"})
        calls = {"n": 0}
        def side_effect(name):
            if name == "ristoranti":
                return ristorante_q
            calls["n"] += 1
            mapping = {1: turni_q, 2: dipendenti_q, 3: storico_q, 4: attivi_q}
            return mapping.get(calls["n"], ex_dipendenti_q)
        ctx, _ = _patch_workspace(side_effect)
        with ctx:
            res = workspace.ws_personale_export_mensile(mese="2026-07", authorization="Bearer x")

        wb = load_workbook(BytesIO(res.body))
        ws_riep = wb["Riepilogo"]
        righe = {r[0]: r for r in ws_riep.iter_rows(min_row=3, values_only=True) if r[0]}
        assert "Luigi" in righe, "l'ex-dipendente deve comparire nel Riepilogo"
        # 8h a 10€/h ciascuno: Mario 80€ + Luigi 80€ = 160€ nel TOTALE.
        assert righe["TOTALE"][7] == 160.0
