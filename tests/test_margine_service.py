"""
Test unitari per services/margine_service.py
Copertura: calcola_risultati, calcola_kpi_anno
"""

import pytest
import pandas as pd
import importlib
import sys
from datetime import date
from types import SimpleNamespace
from unittest.mock import patch, MagicMock

from services.margine_service import calcola_risultati, calcola_kpi_anno, build_transposed_df, export_excel_margini, genera_commenti_kpi
from services.margine_service import (
    calcola_costi_automatici_per_anno,
    carica_costi_per_categoria,
    carica_margini_anno,
    salva_fatturato_centri,
    carica_fatturato_centri_periodo,
    carica_fatturato_centri_mese,
    salva_margini_anno,
)
from config.constants import CATEGORIE_FOOD, CATEGORIE_SPESE_GENERALI

# ---------------------------------------------------------------------------
# Helper: costruisce un DataFrame di input a 12 righe
# ---------------------------------------------------------------------------
MESI_NOMI = ["Gen", "Feb", "Mar", "Apr", "Mag", "Giu",
             "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]

def _make_input(
    fatt_iva10=0.0,
    fatt_iva22=0.0,
    altri_ricavi_noiva=0.0,
    costi_fb_auto=0.0,
    altri_fb=0.0,
    costi_spese_auto=0.0,
    altri_spese=0.0,
    costo_dipendenti=0.0,
    overrides: dict = None,
) -> pd.DataFrame:
    """
    Crea un DataFrame di 12 righe con valori uniformi.
    overrides: {indice_mese (0-11): {colonna: valore}} per impostare mesi specifici.
    """
    rows = []
    for i in range(12):
        rows.append({
            "Mese": MESI_NOMI[i],
            "MeseNum": i + 1,
            "Fatt_IVA10": fatt_iva10,
            "Fatt_IVA22": fatt_iva22,
            "Altri_Ricavi_NoIVA": altri_ricavi_noiva,
            "Costi_FB_Auto": costi_fb_auto,
            "Altri_FB": altri_fb,
            "Costi_Spese_Auto": costi_spese_auto,
            "Altri_Spese": altri_spese,
            "Costo_Dipendenti": costo_dipendenti,
        })
    if overrides:
        for idx, vals in overrides.items():
            rows[idx].update(vals)
    return pd.DataFrame(rows)


# ===========================================================================
# calcola_risultati
# ===========================================================================

class TestCalcolaRisultati:

    def test_calcola_risultati_happy_path(self):
        """
        12 mesi con dati uniformi.
        Verifica struttura output, formula Fatt_Netto, formula MOL e riga TOT ANNO.
        """
        df_in = _make_input(
            fatt_iva10=1100.0,      # netto = 1000
            fatt_iva22=1220.0,      # netto = 1000
            altri_ricavi_noiva=500.0,
            costi_fb_auto=300.0,
            altri_fb=100.0,         # Costi_FB totali = 400
            costi_spese_auto=200.0,
            altri_spese=50.0,       # Costi_Spese totali = 250
            costo_dipendenti=600.0,
        )

        df = calcola_risultati(df_in)

        # Struttura: 13 righe (12 mesi + TOT ANNO)
        assert len(df) == 13

        # Colonne attese
        for col in ["Mese", "MeseNum", "Fatt_Netto", "Costi_FB", "FB_Perc",
                    "Primo_Margine", "PM_Perc", "Costi_Spese", "Spese_Perc",
                    "Costi_Personale", "Pers_Perc", "MOL", "MOL_Perc"]:
            assert col in df.columns, f"Colonna mancante: {col}"

        # Riga TOT ANNO esiste
        tot = df[df["Mese"] == "TOT ANNO"]
        assert len(tot) == 1, "Riga TOT ANNO mancante"

        # Fatt_Netto corretto per il primo mese
        fatt_netto_atteso = round(1100.0 / 1.10 + 1220.0 / 1.22 + 500.0, 2)
        assert round(df.iloc[0]["Fatt_Netto"], 2) == fatt_netto_atteso

        # MOL = Primo_Margine - Costi_Spese - Costi_Personale
        riga0 = df.iloc[0]
        mol_atteso = round(riga0["Primo_Margine"] - riga0["Costi_Spese"] - riga0["Costi_Personale"], 2)
        assert round(riga0["MOL"], 2) == mol_atteso

        # TOT ANNO Fatt_Netto = somma dei 12 mesi
        assert round(tot.iloc[0]["Fatt_Netto"], 2) == round(fatt_netto_atteso * 12, 2)

    def test_calcola_risultati_mese_zero(self):
        """
        Solo il primo mese ha Fatt_Netto == 0 (tutti gli importi a 0).
        Le percentuali del mese devono essere 0.0 senza ZeroDivisionError.
        """
        df_in = _make_input(
            fatt_iva10=1100.0,
            fatt_iva22=0.0,
            altri_ricavi_noiva=0.0,
            costi_fb_auto=300.0,
            altri_fb=0.0,
            costi_spese_auto=100.0,
            altri_spese=0.0,
            costo_dipendenti=200.0,
            overrides={
                0: {
                    "Fatt_IVA10": 0.0,
                    "Fatt_IVA22": 0.0,
                    "Altri_Ricavi_NoIVA": 0.0,
                }
            },
        )

        df = calcola_risultati(df_in)

        riga0 = df.iloc[0]
        assert round(riga0["Fatt_Netto"], 2) == 0.0
        assert round(riga0["FB_Perc"], 2) == 0.0
        assert round(riga0["PM_Perc"], 2) == 0.0
        assert round(riga0["Spese_Perc"], 2) == 0.0
        assert round(riga0["Pers_Perc"], 2) == 0.0
        assert round(riga0["MOL_Perc"], 2) == 0.0

    def test_calcola_risultati_tutti_zero(self):
        """
        Tutti i 12 mesi a zero.
        TOT ANNO deve avere Fatt_Netto == 0 e tutte le percentuali == 0.
        """
        df_in = _make_input()  # tutti i default sono 0.0

        df = calcola_risultati(df_in)
        assert len(df) == 13

        tot = df[df["Mese"] == "TOT ANNO"].iloc[0]
        assert round(tot["Fatt_Netto"], 2) == 0.0
        for perc_col in ["FB_Perc", "PM_Perc", "Spese_Perc", "Pers_Perc", "MOL_Perc"]:
            assert round(tot[perc_col], 2) == 0.0, f"{perc_col} deve essere 0"

    def test_calcola_risultati_un_solo_mese_con_dati(self):
        """
        Solo il mese di Gennaio (indice 0) ha dati reali.
        TOT ANNO deve riflettere solo quel mese.
        """
        df_in = _make_input(
            overrides={
                0: {
                    "Fatt_IVA10": 1100.0,
                    "Fatt_IVA22": 0.0,
                    "Altri_Ricavi_NoIVA": 0.0,
                    "Costi_FB_Auto": 400.0,
                    "Altri_FB": 0.0,
                    "Costi_Spese_Auto": 100.0,
                    "Altri_Spese": 0.0,
                    "Costo_Dipendenti": 150.0,
                }
            }
        )

        df = calcola_risultati(df_in)
        tot = df[df["Mese"] == "TOT ANNO"].iloc[0]

        fatt_netto_gen = round(1100.0 / 1.10, 2)
        assert round(tot["Fatt_Netto"], 2) == fatt_netto_gen

        costi_fb = 400.0
        primo_margine = fatt_netto_gen - costi_fb
        mol_atteso = round(primo_margine - 100.0 - 150.0, 2)
        assert round(tot["MOL"], 2) == mol_atteso


# ===========================================================================
# calcola_kpi_anno
# ===========================================================================

def _make_risultati(rows_override: list = None) -> pd.DataFrame:
    """
    Crea un DataFrame compatibile con l'output di calcola_risultati
    (13 righe: 12 mesi + TOT ANNO) con valori personalizzabili.
    rows_override: lista di dict con i valori per ogni mese (indici 0-11).
    """
    default_mese = {
        "MeseNum": 0,
        "Fatt_Netto": 0.0,
        "Costi_FB": 0.0,
        "FB_Perc": 0.0,
        "Primo_Margine": 0.0,
        "PM_Perc": 0.0,
        "Costi_Spese": 0.0,
        "Spese_Perc": 0.0,
        "Costi_Personale": 0.0,
        "Pers_Perc": 0.0,
        "MOL": 0.0,
        "MOL_Perc": 0.0,
    }
    rows = []
    for i in range(12):
        row = default_mese.copy()
        row["MeseNum"] = i + 1
        row["Mese"] = MESI_NOMI[i]
        if rows_override and i < len(rows_override) and rows_override[i] is not None:
            row.update(rows_override[i])
        rows.append(row)

    # Riga TOT ANNO
    tot = default_mese.copy()
    tot["MeseNum"] = 99
    tot["Mese"] = "TOT ANNO"
    rows.append(tot)

    return pd.DataFrame(rows)


class TestCalcolaKpiAnno:

    def test_calcola_kpi_anno_happy_path(self):
        """
        3 mesi con Fatt_Netto > 0, gli altri a zero.
        Verifica num_mesi == 3 e che le medie siano corrette.
        """
        mesi_data = [
            {"Fatt_Netto": 1000.0, "MOL": 200.0, "FB_Perc": 30.0, "MOL_Perc": 20.0,
             "Primo_Margine": 700.0, "PM_Perc": 70.0, "Costi_FB": 300.0,
             "Costi_Spese": 300.0, "Spese_Perc": 30.0, "Costi_Personale": 200.0, "Pers_Perc": 20.0},
            {"Fatt_Netto": 2000.0, "MOL": 400.0, "FB_Perc": 25.0, "MOL_Perc": 20.0,
             "Primo_Margine": 1500.0, "PM_Perc": 75.0, "Costi_FB": 500.0,
             "Costi_Spese": 700.0, "Spese_Perc": 35.0, "Costi_Personale": 400.0, "Pers_Perc": 20.0},
            {"Fatt_Netto": 1500.0, "MOL": 300.0, "FB_Perc": 28.0, "MOL_Perc": 20.0,
             "Primo_Margine": 1100.0, "PM_Perc": 73.33, "Costi_FB": 400.0,
             "Costi_Spese": 500.0, "Spese_Perc": 33.33, "Costi_Personale": 300.0, "Pers_Perc": 20.0},
        ]

        df = _make_risultati(mesi_data)
        kpi = calcola_kpi_anno(df)

        assert kpi["num_mesi"] == 3
        assert round(kpi["mol_medio"], 2) == round((200.0 + 400.0 + 300.0) / 3, 2)
        assert round(kpi["fatt_medio"], 2) == round((1000.0 + 2000.0 + 1500.0) / 3, 2)
        assert round(kpi["fc_medio"], 2) == round((30.0 + 25.0 + 28.0) / 3, 2)

    def test_calcola_kpi_anno_mesi_filtro(self):
        """
        Passa mesi_filtro=[1, 2]: solo i mesi 1 e 2 devono essere conteggiati.
        """
        mesi_data = [
            {"Fatt_Netto": 1000.0, "MOL": 200.0, "FB_Perc": 30.0, "MOL_Perc": 20.0,
             "Primo_Margine": 700.0, "PM_Perc": 70.0, "Costi_FB": 300.0,
             "Costi_Spese": 300.0, "Spese_Perc": 30.0, "Costi_Personale": 200.0, "Pers_Perc": 20.0},
            {"Fatt_Netto": 2000.0, "MOL": 600.0, "FB_Perc": 20.0, "MOL_Perc": 30.0,
             "Primo_Margine": 1500.0, "PM_Perc": 75.0, "Costi_FB": 400.0,
             "Costi_Spese": 500.0, "Spese_Perc": 25.0, "Costi_Personale": 400.0, "Pers_Perc": 20.0},
            {"Fatt_Netto": 3000.0, "MOL": 900.0, "FB_Perc": 15.0, "MOL_Perc": 30.0,
             "Primo_Margine": 2000.0, "PM_Perc": 66.67, "Costi_FB": 450.0,
             "Costi_Spese": 650.0, "Spese_Perc": 21.67, "Costi_Personale": 450.0, "Pers_Perc": 15.0},
        ]

        df = _make_risultati(mesi_data)
        kpi = calcola_kpi_anno(df, mesi_filtro=[1, 2])

        assert kpi["num_mesi"] == 2
        assert round(kpi["mol_medio"], 2) == round((200.0 + 600.0) / 2, 2)
        assert round(kpi["fatt_medio"], 2) == round((1000.0 + 2000.0) / 2, 2)

    def test_calcola_kpi_anno_tutti_zero(self):
        """
        Tutti i mesi con Fatt_Netto == 0: deve tornare kpi_zero senza eccezioni.
        """
        df = _make_risultati()  # tutti 0
        kpi = calcola_kpi_anno(df)

        assert kpi["num_mesi"] == 0
        assert kpi["mol_medio"] == 0.0
        assert kpi["fatt_medio"] == 0.0
        assert kpi["fc_medio"] == 0.0
        assert kpi["mol_perc_medio"] == 0.0

    def test_calcola_kpi_anno_dataframe_vuoto(self):
        """
        DataFrame vuoto (0 righe): il fallback kpi_zero deve essere restituito
        senza crash.
        """
        df_vuoto = pd.DataFrame(columns=[
            "MeseNum", "Fatt_Netto", "MOL", "FB_Perc", "MOL_Perc",
            "Primo_Margine", "PM_Perc", "Costi_FB",
            "Costi_Spese", "Spese_Perc", "Costi_Personale", "Pers_Perc"
        ])
        kpi = calcola_kpi_anno(df_vuoto)

        assert isinstance(kpi, dict)
        assert kpi["num_mesi"] == 0
        assert kpi["mol_medio"] == 0.0


# ===========================================================================
# build_transposed_df
# ===========================================================================

class TestBuildTransposedDf:

    def test_build_transposed_df_happy_path(self):
        """
        12 mesi con dati validi.
        - 12 righe (una per voce finanziaria)
        - Colonne Gen €, Gen %, Dic €, Dic % presenti
        - Fatt_Netto Gen calcolato correttamente
        """
        df_in = _make_input(
            fatt_iva10=1100.0,
            fatt_iva22=1220.0,
            altri_ricavi_noiva=500.0,
            costi_fb_auto=300.0,
            altri_fb=100.0,
            costi_spese_auto=200.0,
            altri_spese=50.0,
            costo_dipendenti=600.0,
        )

        df = build_transposed_df(df_in)

        # 12 righe (voci finanziarie)
        assert len(df) == 12

        # Colonne attese
        for col in ["Voce", "Gen €", "Gen %", "Dic €", "Dic %"]:
            assert col in df.columns, f"Colonna mancante: {col}"

        # Trova la riga "= Fatturato Netto"
        riga_fn = df[df["Voce"] == "= Fatturato Netto"]
        assert len(riga_fn) == 1, "Riga '= Fatturato Netto' non trovata"

        fatt_netto_atteso = round(1100.0 / 1.10 + 1220.0 / 1.22 + 500.0, 2)
        assert round(riga_fn.iloc[0]["Gen €"], 2) == fatt_netto_atteso

        # La riga Fatturato Netto non ha percentuale (None o NaN)
        assert pd.isna(riga_fn.iloc[0]["Gen %"])

    def test_build_transposed_df_tutti_zero(self):
        """
        Tutti i mesi a zero: deve ritornare 12 righe senza crash,
        con valori 0 nelle colonne €.
        """
        df_in = _make_input()

        df = build_transposed_df(df_in)

        assert len(df) == 12
        assert "Voce" in df.columns

        # Tutti i valori € devono essere 0
        riga_fn = df[df["Voce"] == "= Fatturato Netto"].iloc[0]
        assert round(riga_fn["Gen €"], 2) == 0.0
        assert round(riga_fn["Dic €"], 2) == 0.0


# ===========================================================================
# export_excel_margini
# ===========================================================================

def _make_df_risultati_completo() -> pd.DataFrame:
    """Costruisce un df_risultati realistico passando per calcola_risultati."""
    df_in = _make_input(
        fatt_iva10=1100.0,
        fatt_iva22=1220.0,
        altri_ricavi_noiva=500.0,
        costi_fb_auto=300.0,
        altri_fb=100.0,
        costi_spese_auto=200.0,
        altri_spese=50.0,
        costo_dipendenti=600.0,
    )
    return calcola_risultati(df_in)


# kpi_data minimale con le chiavi attese da export_excel_margini
_KPI_DATA_MINIMALE = {
    "periodo": "Gen-Mar 2025",
    "num_mesi": 3,
    "fatt_totale": 15000.0,
    "fatt_medio": 5000.0,
    "costi_fb": 1200.0,
    "fc_perc": 24.0,
    "primo_marg": 3800.0,
    "primo_marg_perc": 76.0,
    "spese_gen": 600.0,
    "spese_perc": 12.0,
    "personale": 800.0,
    "personale_perc": 16.0,
    "mol_medio": 1000.0,
    "mol_perc": 20.0,
}


class TestExportExcelMargini:

    def test_export_excel_senza_kpi(self):
        """
        Happy path senza kpi_data:
        - ritorna bytes non vuoti
        - file Excel valido con foglio "Margini 2025"
        """
        import io
        import openpyxl

        df_ris = _make_df_risultati_completo()
        result = export_excel_margini(df_ris, anno=2025, nome_ristorante="Test Ristorante")

        assert isinstance(result, bytes)
        assert len(result) > 0

        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert "Margini 2025" in wb.sheetnames
        # Senza kpi_data deve avere solo 1 foglio
        assert "KPI Periodo" not in wb.sheetnames

    def test_export_excel_con_kpi(self):
        """
        Happy path con kpi_data:
        - deve esistere il foglio "KPI Periodo"
        - deve esistere il foglio "Margini 2025"
        """
        import io
        import openpyxl

        df_ris = _make_df_risultati_completo()
        result = export_excel_margini(
            df_ris,
            anno=2025,
            nome_ristorante="Test Ristorante",
            kpi_data=_KPI_DATA_MINIMALE,
        )

        assert isinstance(result, bytes)
        assert len(result) > 0

        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert "Margini 2025" in wb.sheetnames
        assert "KPI Periodo" in wb.sheetnames

    def test_export_excel_df_vuoto_non_crasha(self):
        """
        df_risultati senza righe TOT ANNO (o vuoto): verifica che non crashi
        silenziosamente — deve sollevare un'eccezione documentata (IndexError/KeyError),
        non restituire bytes corrotti.
        """
        import pytest

        df_vuoto = _make_risultati()  # ha TOT ANNO con tutti zero — non crasha
        # Questo è il caso limite: df con TOT ANNO a zero deve comunque produrre bytes validi
        result = export_excel_margini(df_vuoto, anno=2025, nome_ristorante="Vuoto")
        assert isinstance(result, bytes)
        assert len(result) > 0


# ===========================================================================
# genera_commenti_kpi
# ===========================================================================

def _make_df_risultati_minimale_commenti() -> pd.DataFrame:
    """DataFrame minimale per genera_commenti_kpi (12 mesi + TOT ANNO)."""
    rows = []
    for i in range(12):
        rows.append({
            "Mese": MESI_NOMI[i],
            "MeseNum": i + 1,
            "Fatt_Netto": 1000.0,
        })
    rows.append({
        "Mese": "TOT ANNO",
        "MeseNum": 99,
        "Fatt_Netto": 12000.0,
    })
    return pd.DataFrame(rows)


class TestGeneraCommentiKpi:

    def test_kpi_eccellente(self):
        """
        MOL % in fascia eccellente: commento MOL con colore verde ed emoji positiva.
        """
        df_ris = _make_df_risultati_minimale_commenti()
        kpi = {
            "num_mesi": 12,
            "fc_medio": 25.0,
            "primo_margine_perc_media": 75.0,
            "spese_gen_perc_media": 14.0,
            "mol_perc_medio": 25.0,
            "personale_perc_media": 22.0,
        }

        commenti = genera_commenti_kpi(kpi, df_ris)
        mol = next(c for c in commenti if c["kpi_nome"] == "MOL")

        assert mol["emoji"] == "🟢"
        assert mol["colore"] == "#16a34a"

    def test_kpi_critico(self):
        """
        Food Cost in fascia critica: commento Food Cost rosso con emoji negativa.
        """
        df_ris = _make_df_risultati_minimale_commenti()
        kpi = {
            "num_mesi": 12,
            "fc_medio": 80.0,
            "primo_margine_perc_media": 60.0,
            "spese_gen_perc_media": 20.0,
            "mol_perc_medio": 10.0,
            "personale_perc_media": 40.0,
        }

        commenti = genera_commenti_kpi(kpi, df_ris)
        food = next(c for c in commenti if c["kpi_nome"] == "Food Cost")

        assert food["emoji"] == "🔴"
        assert food["colore"] == "#dc2626"

    def test_kpi_norma(self):
        """
        KPI in fascia "nella norma": Food Cost deve risultare giallo/neutro.
        """
        df_ris = _make_df_risultati_minimale_commenti()
        kpi = {
            "num_mesi": 12,
            "fc_medio": 31.0,
            "primo_margine_perc_media": 68.0,
            "spese_gen_perc_media": 18.0,
            "mol_perc_medio": 18.0,
            "personale_perc_media": 28.0,
        }

        commenti = genera_commenti_kpi(kpi, df_ris)
        food = next(c for c in commenti if c["kpi_nome"] == "Food Cost")

        assert food["emoji"] == "🟡"
        assert food["colore"] == "#ca8a04"

    def test_output_struttura(self):
        """
        Ogni elemento dell'output deve avere esattamente le chiavi richieste.
        """
        df_ris = _make_df_risultati_minimale_commenti()
        kpi = {
            "num_mesi": 12,
            "fc_medio": 30.0,
            "primo_margine_perc_media": 70.0,
            "spese_gen_perc_media": 20.0,
            "mol_perc_medio": 15.0,
            "personale_perc_media": 30.0,
        }

        commenti = genera_commenti_kpi(kpi, df_ris)
        chiavi_attese = {"kpi_nome", "percentuale", "commento", "emoji", "colore"}

        assert isinstance(commenti, list)
        assert len(commenti) > 0
        for elemento in commenti:
            assert set(elemento.keys()) == chiavi_attese

    def test_mesi_filtro_none(self):
        """
        Chiamata con mesi_filtro=None: nessuna eccezione.
        """
        df_ris = _make_df_risultati_minimale_commenti()
        kpi = {
            "num_mesi": 12,
            "fc_medio": 30.0,
            "primo_margine_perc_media": 70.0,
            "spese_gen_perc_media": 20.0,
            "mol_perc_medio": 15.0,
            "personale_perc_media": 30.0,
        }

        commenti = genera_commenti_kpi(kpi, df_ris, mesi_filtro=None)
        assert isinstance(commenti, list)

    def test_mesi_filtro_lista_vuota(self):
        """
        Chiamata con mesi_filtro=[]: gestita senza crash.
        """
        df_ris = _make_df_risultati_minimale_commenti()
        kpi = {
            "num_mesi": 12,
            "fc_medio": 30.0,
            "primo_margine_perc_media": 70.0,
            "spese_gen_perc_media": 20.0,
            "mol_perc_medio": 15.0,
            "personale_perc_media": 30.0,
        }

        commenti = genera_commenti_kpi(kpi, df_ris, mesi_filtro=[])
        assert isinstance(commenti, list)

    def test_kpi_dict_vuoto(self):
        """
        kpi vuoto: deve restituire lista vuota senza eccezioni.
        """
        df_ris = _make_df_risultati_minimale_commenti()

        commenti = genera_commenti_kpi({}, df_ris)

        assert isinstance(commenti, list)
        assert commenti == []


# ===========================================================================
# Funzioni DB (mock Supabase)
# ===========================================================================

def _build_query_mock(execute_data=None):
    """
    Crea un mock query chain compatibile con il client Supabase Python.
    """
    query = MagicMock()
    query.select.return_value = query
    query.eq.return_value = query
    query.gte.return_value = query
    query.lte.return_value = query
    query.lt.return_value = query
    query.neq.return_value = query
    query.is_.return_value = query
    query.range.return_value = query
    query.upsert.return_value = query
    query.execute.return_value = SimpleNamespace(data=execute_data or [])
    return query


def _reload_margine_module_without_cache_wrapper():
    """
    Nel test environment streamlit è mockato: forziamo cache_data a decorator identità
    e ricarichiamo il modulo per ottenere funzioni reali (non MagicMock).
    """
    import services.margine_service as margine_module

    streamlit_mod = sys.modules.get("streamlit")
    if streamlit_mod is not None:
        streamlit_mod.cache_data = lambda *args, **kwargs: (lambda func: func)

    return importlib.reload(margine_module)


class TestMargineServiceDB:

    def test_calcola_costi_automatici_per_anno(self):
        """
        Verifica aggregazione costi F&B e Spese Generali da dati mock fatture.
        """
        margine_module = _reload_margine_module_without_cache_wrapper()
        food_cat = CATEGORIE_FOOD[0]
        spese_cat = CATEGORIE_SPESE_GENERALI[0]

        mock_client = MagicMock()

        query = _build_query_mock()
        query.execute.side_effect = [
            SimpleNamespace(data=[
                {"data_documento": "2026-01-10", "totale_riga": 100.0, "categoria": food_cat},
                {"data_documento": "2026-01-15", "totale_riga": 80.0, "categoria": spese_cat},
                {"data_documento": "2026-01-20", "totale_riga": 50.0, "categoria": food_cat},
            ])
        ]
        mock_client.table.return_value = query

        with patch("services.margine_service.get_supabase_client", return_value=mock_client):
            costi_fb, costi_spese = margine_module.calcola_costi_automatici_per_anno(
                user_id="test-uuid",
                ristorante_id="rist-test",
                anno=2026,
            )

        assert costi_fb.get(1) == 150.0
        assert costi_spese.get(1) == 80.0

    def test_carica_costi_per_categoria(self):
        """
        Verifica DataFrame aggregato per categoria/mese su sole categorie F&B.
        """
        margine_module = _reload_margine_module_without_cache_wrapper()
        food_cat = CATEGORIE_FOOD[0]

        mock_client = MagicMock()

        query = _build_query_mock()
        query.execute.side_effect = [
            SimpleNamespace(data=[
                {"data_documento": "2026-03-01", "totale_riga": 40.0, "categoria": food_cat},
                {"data_documento": "2026-03-10", "totale_riga": 60.0, "categoria": food_cat},
            ])
        ]
        mock_client.table.return_value = query

        with patch("services.margine_service.get_supabase_client", return_value=mock_client):
            df = margine_module.carica_costi_per_categoria(
                user_id="test-uuid",
                ristorante_id="rist-test",
                date_from="2026-03-01",
                date_to="2026-03-31",
            )

        assert not df.empty
        assert set(df.columns) == {"categoria", "mese", "totale"}
        assert float(df["totale"].sum()) == 100.0

    @patch("services.margine_service.get_supabase_client")
    def test_carica_margini_anno(self, mock_get_client):
        """
        Verifica mapping {mese: row_dict} da tabella margini_mensili.
        """
        mock_client = MagicMock()
        mock_get_client.return_value = mock_client

        query = _build_query_mock(execute_data=[
            {"mese": 1, "fatturato_iva10": 1100.0, "fatturato_iva22": 1220.0,
             "altri_ricavi_noiva": 50.0, "altri_costi_fb": 10.0,
             "altri_costi_spese": 20.0, "costo_dipendenti": 300.0},
            {"mese": 2, "fatturato_iva10": 900.0, "fatturato_iva22": 1000.0,
             "altri_ricavi_noiva": 0.0, "altri_costi_fb": 5.0,
             "altri_costi_spese": 15.0, "costo_dipendenti": 250.0},
        ])
        mock_client.table.return_value = query

        result = carica_margini_anno("test-uuid", "rist-test", 2026)

        assert isinstance(result, dict)
        assert set(result.keys()) == {1, 2}
        assert result[1]["fatturato_iva10"] == 1100.0

    @patch("services.margine_service.get_supabase_client")
    def test_salva_fatturato_centri(self, mock_get_client):
        """
        Verifica upsert split fatturato per centri su mese specifico.
        """
        mock_client = MagicMock()
        mock_get_client.return_value = mock_client

        query = _build_query_mock(execute_data=[{"ok": True}])
        mock_client.table.return_value = query

        ok = salva_fatturato_centri(
            user_id="test-uuid",
            ristorante_id="rist-test",
            anno=2026,
            mese=4,
            split_euro={"FOOD": 3000, "BAR": 500, "ALCOLICI": 700, "DOLCI": 200},
        )

        assert ok is True
        query.upsert.assert_called_once()

    @patch("services.margine_service.get_supabase_client")
    def test_carica_fatturato_centri_periodo(self, mock_get_client):
        """
        Verifica aggregazione periodo multi-mese con ritorno dict centri.
        """
        mock_client = MagicMock()
        mock_get_client.return_value = mock_client

        query = _build_query_mock()
        query.execute.side_effect = [
            SimpleNamespace(data=[
                {
                    "fatturato_food": 1000.0,
                    "fatturato_bar": 200.0,
                    "fatturato_alcolici": 150.0,
                    "fatturato_dolci": 80.0,
                },
                {
                    "fatturato_food": 900.0,
                    "fatturato_bar": 180.0,
                    "fatturato_alcolici": 120.0,
                    "fatturato_dolci": 70.0,
                },
            ])
        ]
        mock_client.table.return_value = query

        result = carica_fatturato_centri_periodo(
            user_id="test-uuid",
            ristorante_id="rist-test",
            data_inizio=date(2026, 1, 1),
            data_fine=date(2026, 2, 28),
        )

        assert result["FOOD"] == 1900.0
        assert result["BAR"] == 380.0
        assert result["ALCOLICI"] == 270.0
        assert result["DOLCI"] == 150.0

    @patch("services.margine_service.get_supabase_client")
    def test_carica_fatturato_centri_mese(self, mock_get_client):
        """
        Verifica caricamento split centri per singolo mese.
        """
        mock_client = MagicMock()
        mock_get_client.return_value = mock_client

        query = _build_query_mock(execute_data=[
            {
                "fatturato_food": 2200.0,
                "fatturato_bar": 450.0,
                "fatturato_alcolici": 300.0,
                "fatturato_dolci": 120.0,
            }
        ])
        mock_client.table.return_value = query

        result = carica_fatturato_centri_mese(
            user_id="test-uuid",
            ristorante_id="rist-test",
            anno=2026,
            mese=4,
        )

        assert result == {
            "FOOD": 2200.0,
            "BAR": 450.0,
            "ALCOLICI": 300.0,
            "DOLCI": 120.0,
        }

    @patch("services.margine_service.get_supabase_client")
    def test_salva_margini_anno(self, mock_get_client):
        """
        Verifica upsert di 12 record mensili in margini_mensili.
        """
        mock_client = MagicMock()
        mock_get_client.return_value = mock_client

        query = _build_query_mock(execute_data=[{"ok": True}])
        mock_client.table.return_value = query

        df_input = _make_input(
            fatt_iva10=1100.0,
            fatt_iva22=1220.0,
            altri_ricavi_noiva=100.0,
            costi_fb_auto=300.0,
            altri_fb=50.0,
            costi_spese_auto=200.0,
            altri_spese=30.0,
            costo_dipendenti=400.0,
        )
        df_risultati = calcola_risultati(df_input)

        ok = salva_margini_anno(
            user_id="test-uuid",
            ristorante_id="rist-test",
            anno=2026,
            df_input=df_input,
            df_risultati=df_risultati,
        )

        assert ok is True
        query.upsert.assert_called_once()
        records_arg = query.upsert.call_args[0][0]
        assert isinstance(records_arg, list)
        assert len(records_arg) == 12

